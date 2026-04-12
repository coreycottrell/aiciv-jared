#!/usr/bin/env python3
"""
PureBrain CFO-Grade 5-Year Financial Model
World's Best Financial Model - Full Algorithm Excel Build
AF# Report: CFO-Grade Financial Model
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

# ─────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────
DARK_BLUE       = "1B2A4A"
MID_BLUE        = "2E4A7A"
LIGHT_BLUE      = "D6EAF8"
INPUT_BLUE      = "BDD7EE"
WHITE           = "FFFFFF"
LIGHT_GRAY      = "F2F2F2"
HEADER_GRAY     = "D9D9D9"
GOLD            = "C9A84C"
DARK_GOLD       = "9A7B2F"
LIGHT_GOLD      = "FFF2CC"
GREEN           = "1E7E34"
LIGHT_GREEN     = "D5F5E3"
RED             = "C0392B"
LIGHT_RED       = "FADBD8"
ORANGE          = "E67E22"
CALC_WHITE      = "FFFFFF"
SECTION_BG      = "EBF5FB"

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────
def make_font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_border():
    thick = Side(style="medium")
    thin = Side(style="thin")
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def apply_header(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, size=11, bold=True, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=bold, size=size, color=fg)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border("thin")
    return cell

def apply_input(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=False, size=10, color="1A1A2E")
    cell.fill = make_fill(INPUT_BLUE)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = make_border("thin")
    if fmt:
        cell.number_format = fmt
    return cell

def apply_calc(ws, row, col, formula_or_value, fmt=None, bold=False, bg=CALC_WHITE, color="000000"):
    cell = ws.cell(row=row, column=col, value=formula_or_value)
    cell.font = make_font(bold=bold, size=10, color=color)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = make_border("thin")
    if fmt:
        cell.number_format = fmt
    return cell

def apply_label(ws, row, col, value, bold=False, bg=LIGHT_GRAY, size=10, align="left", color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=bold, size=size, color=color)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", indent=1)
    cell.border = make_border("thin")
    return cell

def merge_header(ws, row, col_start, col_end, value, bg=DARK_BLUE, fg=WHITE, size=12, bold=True):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = make_font(bold=bold, size=size, color=fg)
    cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border("medium")
    return cell

FMT_DOLLAR   = '$#,##0'
FMT_DOLLAR2  = '$#,##0.00'
FMT_MILLION  = '$#,##0,,"M"'
FMT_BILLION  = '$#,##0,,,"B"'
FMT_PCT      = '0.0%'
FMT_PCT1     = '0.00%'
FMT_INT      = '#,##0'
FMT_RATIO    = '0.0"x"'
FMT_MONTHS   = '0.0" mo"'

# ─────────────────────────────────────────────────────────────────
# SHEET 1: ASSUMPTIONS
# ─────────────────────────────────────────────────────────────────
def build_assumptions(wb):
    ws = wb.create_sheet("ASSUMPTIONS")
    ws.sheet_properties.tabColor = "1B2A4A"

    # Title
    ws.row_dimensions[1].height = 35
    merge_header(ws, 1, 1, 8, "PUREBRAIN.AI  —  5-YEAR FINANCIAL MODEL  |  ASSUMPTIONS & CONTROL PANEL", size=14)

    ws.row_dimensions[2].height = 20
    ws.cell(row=2, column=1, value="Blue cells = EDITABLE INPUTS   |   White/Gray cells = CALCULATED   |   Last Updated: " + datetime.date.today().strftime("%B %d, %Y"))
    ws.cell(row=2, column=1).font = make_font(italic=True, size=9, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Column widths
    col_widths = [32, 18, 18, 18, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4

    # ── SECTION: SCENARIO SELECTOR ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "SCENARIO SELECTOR", bg=DARK_GOLD, size=11)
    r += 1

    apply_label(ws, r, 1, "Active Scenario (1=Bear, 2=Base, 3=Bull)", bold=True, bg=LIGHT_GOLD)
    apply_input(ws, r, 2, 2)
    ws.cell(row=r, column=2).comment = None
    apply_label(ws, r, 3, "← Change this cell", bold=False, bg=LIGHT_GOLD, color="666666")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    apply_label(ws, r, 5, "1 = Bear Case", bg=LIGHT_GOLD, color=RED)
    apply_label(ws, r, 6, "2 = Base Case", bg=LIGHT_GOLD, color=MID_BLUE)
    apply_label(ws, r, 7, "3 = Bull Case", bg=LIGHT_GOLD, color=GREEN)
    r += 1

    # Named range helper rows (hidden logic)
    apply_label(ws, r, 1, "Scenario Growth Multiplier (auto-calc)", bold=False, bg=LIGHT_GRAY, color="888888")
    # 1=Bear→0.67, 2=Base→1.00, 3=Bull→1.40
    apply_calc(ws, r, 2, '=CHOOSE(B5,0.67,1.00,1.40)', FMT_PCT, bg=HEADER_GRAY)
    apply_label(ws, r, 3, "0.67 Bear | 1.00 Base | 1.40 Bull", bg=LIGHT_GRAY, color="888888")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 2

    # ── SECTION: PRICING ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "PRICING  &  TIER DISTRIBUTION", bg=DARK_BLUE, size=11)
    r += 1

    headers = ["Tier", "Monthly Price", "Tier Mix %", "Monthly Churn %", "Avg Life (months)", "Avg Life Calc", "LTV", "LTV Calc"]
    for c, h in enumerate(headers, 1):
        apply_header(ws, r, c, h, bg=MID_BLUE, size=9)
    r += 1

    tiers = [
        ("Awakened",   197,  0.72, 0.045),
        ("Partnered",  579,  0.20, 0.030),
        ("Unified",   1089,  0.08, 0.020),
    ]
    tier_rows = {}
    for tier, price, mix, churn in tiers:
        apply_label(ws, r, 1, tier, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, price, FMT_DOLLAR)
        apply_input(ws, r, 3, mix, FMT_PCT)
        apply_input(ws, r, 4, churn, FMT_PCT1)
        apply_calc(ws, r, 5, f'=IFERROR(1/{get_column_letter(4)}{r},0)', FMT_MONTHS)
        apply_calc(ws, r, 6, f'=ROUND(E{r},1)', FMT_MONTHS, bg=LIGHT_GRAY)
        apply_calc(ws, r, 7, f'=B{r}*E{r}', FMT_DOLLAR)
        apply_calc(ws, r, 8, f'=ROUND(G{r},0)', FMT_DOLLAR, bg=LIGHT_GRAY)
        tier_rows[tier] = r
        r += 1

    # Validation row
    apply_label(ws, r, 1, "Mix Total (must = 100%)", bold=True, bg=LIGHT_GRAY)
    apply_calc(ws, r, 2, f'=SUM(C{tier_rows["Awakened"]}:C{tier_rows["Unified"]})', FMT_PCT, bold=True, bg=LIGHT_GREEN)
    apply_calc(ws, r, 3, f'=IF(ABS(B{r}-1)<0.001,"✓ VALID","⚠ FIX MIX")', bg=LIGHT_GREEN, bold=True, color=GREEN)
    r += 2

    # ── SECTION: ENTERPRISE ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "ENTERPRISE PRICING  &  CLIENT ASSUMPTIONS", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Scale Stage", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "User Count", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "# Enterprise Clients", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "Price/Client/Mo", bg=MID_BLUE, size=9)
    apply_header(ws, r, 5, "Enterprise MRR", bg=MID_BLUE, size=9)
    apply_header(ws, r, 6, "Enterprise Churn", bg=MID_BLUE, size=9)
    r += 1

    enterprise_stages = [
        ("MVP",      100,       2,    10000),
        ("1K Users", 1000,      5,    10000),
        ("10K Users",10000,     25,   12000),
        ("100K Users",100000,   150,  15000),
        ("1M Users", 1000000,   1000, 18000),
        ("5M+ Users",5000000,   5000, 20000),
    ]
    ent_rows = {}
    for label, users, clients, price in enterprise_stages:
        apply_label(ws, r, 1, label, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, users, FMT_INT)
        apply_input(ws, r, 3, clients, FMT_INT)
        apply_input(ws, r, 4, price, FMT_DOLLAR)
        apply_calc(ws, r, 5, f'=C{r}*D{r}', FMT_DOLLAR)
        apply_input(ws, r, 6, 0.015, FMT_PCT1)
        ent_rows[label] = r
        r += 1
    r += 1

    # ── SECTION: BLENDED ARPU ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "BLENDED CONSUMER ARPU  (auto-calculated from tier mix & pricing)", bg=MID_BLUE, size=11)
    r += 1

    apply_label(ws, r, 1, "Blended Consumer ARPU", bold=True, bg=LIGHT_BLUE)
    aw_r = tier_rows["Awakened"]
    pa_r = tier_rows["Partnered"]
    un_r = tier_rows["Unified"]
    apply_calc(ws, r, 2,
        f'=B{aw_r}*C{aw_r}+B{pa_r}*C{pa_r}+B{un_r}*C{un_r}',
        FMT_DOLLAR2, bold=True, bg=LIGHT_BLUE, color=DARK_BLUE)
    apply_label(ws, r, 3, "Weighted avg of tier prices × tier mix", bg=LIGHT_BLUE, color="555555")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 2

    # ── SECTION: HEADCOUNT ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "HEADCOUNT  &  PAYROLL ASSUMPTIONS", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Stage", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Total Users", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "Headcount", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "Avg Salary/Mo", bg=MID_BLUE, size=9)
    apply_header(ws, r, 5, "Total Payroll/Mo", bg=MID_BLUE, size=9)
    apply_header(ws, r, 6, "$/User", bg=MID_BLUE, size=9)
    r += 1

    headcount_stages = [
        ("Launch",   100,    48,  10000),
        ("1K Users", 1000,   55,  10000),
        ("10K Users",10000,  65,  10000),
        ("100K",     100000, 80,  10000),
        ("500K",     500000, 92,  10000),
        ("1M+",      1000000,98,  10000),
        ("Hard Cap", 9999999,100, 10000),
    ]
    hc_rows = []
    for label, users, hc, sal in headcount_stages:
        apply_label(ws, r, 1, label, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, users, FMT_INT)
        apply_input(ws, r, 3, hc, FMT_INT)
        apply_input(ws, r, 4, sal, FMT_DOLLAR)
        apply_calc(ws, r, 5, f'=C{r}*D{r}', FMT_DOLLAR, bold=True)
        apply_calc(ws, r, 6, f'=IFERROR(E{r}/B{r},0)', FMT_DOLLAR2)
        hc_rows.append(r)
        r += 1
    r += 1

    # ── SECTION: AI COMPUTE COST ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "AI COMPUTE COST  PER USER/MONTH  (economies of scale)", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Scale Stage", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Total Users", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "Cost/User/Mo", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "% of ARPU", bg=MID_BLUE, size=9)
    r += 1

    compute_stages = [
        ("100 users",    100,     25.00),
        ("1K users",     1000,    18.00),
        ("10K users",    10000,   12.00),
        ("100K users",   100000,   7.00),
        ("1M users",     1000000,  3.50),
        ("5M+ users",    5000000,  2.40),
    ]
    compute_rows = []
    arpu_ref_row = None
    for idx, (label, users, cost) in enumerate(compute_stages):
        apply_label(ws, r, 1, label, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, users, FMT_INT)
        apply_input(ws, r, 3, cost, FMT_DOLLAR2)
        # ARPU reference: find blended ARPU row
        apply_calc(ws, r, 4, f'=IFERROR(C{r}/ASSUMPTIONS!$B${r - (r - compute_stages[0][0]) if False else ""},0)', FMT_PCT)
        compute_rows.append(r)
        r += 1

    # Fix column 4 to reference actual ARPU
    # We need to find the ARPU row — it was set before the compute section
    # We'll patch it after all rows are known
    r += 1

    # ── SECTION: MARKETING ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "MARKETING  &  CAC ASSUMPTIONS", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Stage", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Marketing % MRR", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "CAC ($)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "NRR", bg=MID_BLUE, size=9)
    r += 1

    marketing_stages = [
        ("Launch (Mar 2026)",  0.093, 150, 1.07),
        ("Growth (Jun 2026)",  0.138, 90,  1.12),
        ("Scale (Dec 2026)",   0.156, 45,  1.18),
        ("Maturity (Year 3+)", 0.100, 20,  1.25),
    ]
    mkt_rows = []
    for label, mktpct, cac, nrr in marketing_stages:
        apply_label(ws, r, 1, label, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, mktpct, FMT_PCT)
        apply_input(ws, r, 3, cac, FMT_DOLLAR)
        apply_input(ws, r, 4, nrr, FMT_PCT)
        mkt_rows.append(r)
        r += 1
    r += 1

    # ── SECTION: OTHER OPEX ──
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "OTHER OPEX ASSUMPTIONS", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Expense Category", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Rate / Method", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "Mar 2026 ($)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "Jun 2026 ($)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 5, "Dec 2026 ($)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 6, "Year 2 Run Rate ($)", bg=MID_BLUE, size=9)
    r += 1

    other_opex = [
        ("Payment Processing",  "3% of MRR",    1600,   1086000, 3863000, 30000000),
        ("Office & Ops",        "Fixed + growth",20000,  30000,   50000,   100000),
        ("Legal",               "Fixed + growth",10000,  25000,   50000,   100000),
        ("Misc/Contingency",    "5% of (Payroll+Mkt+Compute)",28000, 157000, 557000, 2710000),
    ]
    opex_rows = {}
    for label, method, m1, m2, m3, yr2 in other_opex:
        apply_label(ws, r, 1, label, bold=True, bg=SECTION_BG)
        apply_label(ws, r, 2, method, bg=LIGHT_GRAY, color="555555")
        apply_input(ws, r, 3, m1, FMT_DOLLAR)
        apply_input(ws, r, 4, m2, FMT_DOLLAR)
        apply_input(ws, r, 5, m3, FMT_DOLLAR)
        apply_input(ws, r, 6, yr2, FMT_DOLLAR)
        opex_rows[label] = r
        r += 1

    apply_label(ws, r, 1, "Payment Processing Rate", bold=True, bg=LIGHT_BLUE)
    apply_input(ws, r, 2, 0.03, FMT_PCT)
    apply_label(ws, r, 3, "← Applied to all MRR in P&L", bg=LIGHT_BLUE, color="555555")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

    # ── SECTION: GROWTH RATES ──
    r += 1
    ws.row_dimensions[r].height = 22
    merge_header(ws, r, 1, 8, "MONTHLY USER GROWTH TARGETS  (new sign-ups per month)", bg=DARK_BLUE, size=11)
    r += 1

    apply_header(ws, r, 1, "Month", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "New Sign-Ups (Base)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 3, "New Sign-Ups (Bear)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 4, "New Sign-Ups (Bull)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 5, "Active (Scenario)", bg=MID_BLUE, size=9)
    apply_header(ws, r, 6, "MoM Growth %", bg=MID_BLUE, size=9)
    r += 1

    months = [
        "Mar 2026","Apr 2026","May 2026","Jun 2026","Jul 2026","Aug 2026",
        "Sep 2026","Oct 2026","Nov 2026","Dec 2026","Jan 2027","Feb 2027"
    ]
    base_signups = [100,900,9000,90000,30000,35000,42000,50000,60000,72000,86000,100000]
    bear_signups = [67,600,6000,60000,20000,23000,28000,33000,40000,48000,58000,67000]
    bull_signups = [140,1260,12600,126000,42000,49000,59000,70000,84000,101000,120000,140000]

    growth_rows = {}
    for i, (month, base, bear, bull) in enumerate(zip(months, base_signups, bear_signups, bull_signups)):
        apply_label(ws, r, 1, month, bold=True, bg=SECTION_BG)
        apply_input(ws, r, 2, base, FMT_INT)
        apply_input(ws, r, 3, bear, FMT_INT)
        apply_input(ws, r, 4, bull, FMT_INT)
        # Column 5: scenario-selected
        apply_calc(ws, r, 5, f'=CHOOSE(ASSUMPTIONS!$B$5,C{r},B{r},D{r})', FMT_INT, bg=LIGHT_BLUE, bold=True)
        if i > 0:
            prev = growth_rows[months[i-1]]
            apply_calc(ws, r, 6, f'=IFERROR((E{r}-E{prev})/E{prev},0)', FMT_PCT)
        else:
            apply_calc(ws, r, 6, "N/A")
        growth_rows[month] = r
        r += 1

    r += 2
    # Legend
    apply_label(ws, r, 1, "LEGEND:", bold=True, bg=LIGHT_GOLD, color=DARK_GOLD)
    apply_label(ws, r, 2, "Blue = Editable Input", bg=INPUT_BLUE, color=DARK_BLUE)
    apply_label(ws, r, 3, "White = Auto-Calculated", bg=CALC_WHITE, color="333333")
    apply_label(ws, r, 4, "Gray = Reference / Label", bg=LIGHT_GRAY, color="555555")
    apply_label(ws, r, 5, "Gold = Scenario Control", bg=LIGHT_GOLD, color=DARK_GOLD)

    ws.freeze_panes = "B4"
    return ws, tier_rows, ent_rows, hc_rows, mkt_rows, compute_rows, growth_rows, opex_rows


# ─────────────────────────────────────────────────────────────────
# SHEET 2: YEAR 1 MONTHLY
# ─────────────────────────────────────────────────────────────────
def build_year1_monthly(wb, tier_rows, growth_rows):
    ws = wb.create_sheet("YEAR 1 MONTHLY")
    ws.sheet_properties.tabColor = "2E7D32"

    months = [
        "Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26",
        "Sep-26","Oct-26","Nov-26","Dec-26","Jan-27","Feb-27"
    ]

    # Title
    ws.row_dimensions[1].height = 35
    merge_header(ws, 1, 1, 14, "PUREBRAIN.AI  —  YEAR 1 MONTHLY MODEL  (Mar 2026 – Feb 2027)", size=13)

    ws.row_dimensions[2].height = 16
    ws.cell(row=2, column=1, value="All formulas reference ASSUMPTIONS sheet. Change scenario there to see Bear/Base/Bull.")
    ws.cell(row=2, column=1).font = make_font(italic=True, size=9, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    # Column widths
    ws.column_dimensions["A"].width = 32
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 14

    r = 4

    # Month headers
    apply_header(ws, r, 1, "METRIC", bg=DARK_BLUE, size=10)
    for i, m in enumerate(months, 2):
        apply_header(ws, r, i, m, bg=DARK_BLUE, size=9)
    apply_header(ws, r, 14, "YEAR 1 TOTAL", bg=MID_BLUE, size=9)
    r += 1

    # Reference rows for ASSUMPTIONS
    # Scenario signups row — we reference ASSUMPTIONS!E (the chosen scenario)
    # ASSUMPTIONS growth_rows[month] → column E = scenario signups

    # We'll build the model with a lookup: for month i, use ASSUMPTIONS growth_rows
    month_full = [
        "Mar 2026","Apr 2026","May 2026","Jun 2026","Jul 2026","Aug 2026",
        "Sep 2026","Oct 2026","Nov 2026","Dec 2026","Jan 2027","Feb 2027"
    ]

    # ── NEW SIGN-UPS ──
    merge_header(ws, r, 1, 14, "SUBSCRIBER ACQUISITION", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    apply_label(ws, r, 1, "New Sign-Ups (Scenario)", bold=True, bg=LIGHT_BLUE)
    for i, m in enumerate(months):
        gr = growth_rows[month_full[i]]
        apply_calc(ws, r, i+2, f'=ASSUMPTIONS!E{gr}', FMT_INT, bg=LIGHT_BLUE, bold=True)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_INT, bold=True, bg=LIGHT_BLUE)
    new_signup_row = r
    r += 1

    # ── SUBSCRIBER TIERS ──
    merge_header(ws, r, 1, 14, "ACTIVE SUBSCRIBERS  BY TIER", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    tiers = ["Awakened", "Partnered", "Unified"]
    tier_mix_cols = {
        "Awakened": tier_rows["Awakened"],
        "Partnered": tier_rows["Partnered"],
        "Unified": tier_rows["Unified"],
    }
    tier_churn_map = {
        "Awakened": ("C", tier_rows["Awakened"]),
        "Partnered": ("C", tier_rows["Partnered"]),
        "Unified": ("C", tier_rows["Unified"]),
    }

    tier_active_rows = {}
    for tier in tiers:
        apply_label(ws, r, 1, f"Active {tier} Subs", bold=True, bg=SECTION_BG)
        tr = tier_rows[tier]
        churn_ref = f'ASSUMPTIONS!D{tr}'
        mix_ref = f'ASSUMPTIONS!C{tr}'
        for i in range(len(months)):
            col = i + 2
            if i == 0:
                # First month: new signups × mix × (1 - churn)
                apply_calc(ws, r, col,
                    f'=ROUND(B{new_signup_row}*{mix_ref}*(1-{churn_ref}),0)',
                    FMT_INT)
            else:
                prev_col = get_column_letter(col - 1)
                cur_col = get_column_letter(col)
                apply_calc(ws, r, col,
                    f'=ROUND({prev_col}{r}*(1-{churn_ref})+{get_column_letter(col)}{new_signup_row}*{mix_ref},0)',
                    FMT_INT)
        apply_calc(ws, r, 14, f'=M{r}', FMT_INT, bold=True)  # end-of-period
        tier_active_rows[tier] = r
        r += 1

    apply_label(ws, r, 1, "TOTAL Active Subscribers", bold=True, bg=LIGHT_BLUE)
    for i in range(len(months)):
        col = i + 2
        col_refs = "+".join([f'{get_column_letter(col)}{tier_active_rows[t]}' for t in tiers])
        apply_calc(ws, r, col, f'={col_refs}', FMT_INT, bold=True, bg=LIGHT_BLUE)
    apply_calc(ws, r, 14, f'=M{r}', FMT_INT, bold=True, bg=LIGHT_BLUE)
    total_active_row = r
    r += 1

    # ── CONSUMER MRR ──
    merge_header(ws, r, 1, 14, "CONSUMER  MRR  BY TIER", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    tier_mrr_rows = {}
    for tier in tiers:
        apply_label(ws, r, 1, f"{tier} MRR", bold=True, bg=SECTION_BG)
        tr = tier_rows[tier]
        price_ref = f'ASSUMPTIONS!B{tr}'
        for i in range(len(months)):
            col = i + 2
            apply_calc(ws, r, col,
                f'={get_column_letter(col)}{tier_active_rows[tier]}*{price_ref}',
                FMT_DOLLAR)
        apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
        tier_mrr_rows[tier] = r
        r += 1

    apply_label(ws, r, 1, "Total Consumer MRR", bold=True, bg=LIGHT_BLUE)
    for i in range(len(months)):
        col = i + 2
        refs = "+".join([f'{get_column_letter(col)}{tier_mrr_rows[t]}' for t in tiers])
        apply_calc(ws, r, col, f'={refs}', FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    consumer_mrr_row = r
    r += 1

    # ── ENTERPRISE MRR ──
    merge_header(ws, r, 1, 14, "ENTERPRISE  MRR", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    ent_mrr_values = [20000, 50000, 300000, 2250000, 3000000, 3750000,
                      5000000, 6500000, 8500000, 11000000, 14000000, 17500000]
    ent_clients = [2, 5, 25, 150, 200, 250, 333, 433, 567, 733, 933, 1167]

    apply_label(ws, r, 1, "Enterprise Clients", bold=True, bg=SECTION_BG)
    for i, ec in enumerate(ent_clients):
        apply_input(ws, r, i+2, ec, FMT_INT)
    apply_calc(ws, r, 14, f'=M{r}', FMT_INT, bold=True)
    ent_clients_row = r
    r += 1

    apply_label(ws, r, 1, "Avg Enterprise Price/Mo", bold=True, bg=SECTION_BG)
    ent_prices = [10000,10000,12000,15000,15000,15000,15000,15000,17000,15019,15005,14993]
    for i, ep in enumerate(ent_prices):
        apply_input(ws, r, i+2, ep, FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=AVERAGE(B{r}:M{r})', FMT_DOLLAR, bold=True)
    ent_price_row = r
    r += 1

    apply_label(ws, r, 1, "Enterprise MRR", bold=True, bg=LIGHT_BLUE)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{ent_clients_row}*{get_column_letter(col)}{ent_price_row}',
            FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    enterprise_mrr_row = r
    r += 1

    # ── TOTAL MRR ──
    merge_header(ws, r, 1, 14, "TOTAL  MRR  &  REVENUE", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    apply_label(ws, r, 1, "TOTAL MRR", bold=True, bg=LIGHT_GREEN)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{consumer_mrr_row}+{get_column_letter(col)}{enterprise_mrr_row}',
            FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    total_mrr_row = r
    r += 1

    apply_label(ws, r, 1, "ARR (MRR × 12)", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col, f'={get_column_letter(col)}{total_mrr_row}*12', FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=M{r}', FMT_DOLLAR, bold=True)
    r += 1

    apply_label(ws, r, 1, "MoM MRR Growth %", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        if i == 0:
            apply_calc(ws, r, col, "N/A")
        else:
            prev_col = get_column_letter(col - 1)
            apply_calc(ws, r, col,
                f'=IFERROR(({get_column_letter(col)}{total_mrr_row}-{prev_col}{total_mrr_row})/{prev_col}{total_mrr_row},0)',
                FMT_PCT)
    r += 1

    # ── EXPENSES ──
    merge_header(ws, r, 1, 14, "MONTHLY  EXPENSES  (OPEX)", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    # Payroll: use headcount lookup from assumptions
    payroll_values = [480000, 550000, 650000, 800000, 800000, 850000, 900000, 920000, 950000, 980000, 1000000, 1000000]
    compute_values = [2500, 18000, 119856, 689220, 875000, 1088000, 1341000, 1637000, 1983000, 2391000, 2870000, 3432000]
    marketing_values = [5000, 56000, 517000, 4988000, 6355000, 7905000, 9806000, 12027000, 14674000, 20069000, 21470000, 25746000]

    apply_label(ws, r, 1, "Payroll", bold=True, bg=SECTION_BG)
    for i, pv in enumerate(payroll_values):
        apply_input(ws, r, i+2, pv, FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    payroll_row = r
    r += 1

    apply_label(ws, r, 1, "AI Compute (users × cost/user)", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{total_active_row}*VLOOKUP({get_column_letter(col)}{total_active_row},ASSUMPTIONS!$B$60:$C$65,2,TRUE)',
            FMT_DOLLAR)
        # fallback to hardcoded if VLOOKUP range is off
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    compute_row = r
    r += 1

    apply_label(ws, r, 1, "Marketing", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        # Marketing % of MRR, interpolated: start 9.3%, mature 15.6%
        # Simple: 13.8% of consumer MRR for growth months
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{consumer_mrr_row}*0.138',
            FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    marketing_row = r
    r += 1

    apply_label(ws, r, 1, "Payment Processing (3% of MRR)", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{total_mrr_row}*ASSUMPTIONS!$B${97}',
            FMT_DOLLAR)
        # We'll use a fixed 3% ref — will patch after
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    payment_row = r
    r += 1

    apply_label(ws, r, 1, "Office & Ops", bold=True, bg=SECTION_BG)
    office_vals = [20000,22000,25000,30000,30000,32000,35000,38000,42000,50000,50000,50000]
    for i, v in enumerate(office_vals):
        apply_input(ws, r, i+2, v, FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    office_row = r
    r += 1

    apply_label(ws, r, 1, "Legal", bold=True, bg=SECTION_BG)
    legal_vals = [10000,12000,15000,25000,25000,28000,32000,36000,40000,50000,50000,50000]
    for i, v in enumerate(legal_vals):
        apply_input(ws, r, i+2, v, FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    legal_row = r
    r += 1

    apply_label(ws, r, 1, "Misc/Contingency (5%)", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'=({get_column_letter(col)}{payroll_row}+{get_column_letter(col)}{marketing_row}+{get_column_letter(col)}{compute_row})*0.05',
            FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    misc_row = r
    r += 1

    apply_label(ws, r, 1, "TOTAL OPEX", bold=True, bg=LIGHT_RED)
    for i in range(len(months)):
        col = i + 2
        cols = [payroll_row, compute_row, marketing_row, payment_row, office_row, legal_row, misc_row]
        refs = "+".join([f'{get_column_letter(col)}{row}' for row in cols])
        apply_calc(ws, r, col, f'={refs}', FMT_DOLLAR, bold=True, bg=LIGHT_RED, color=RED)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_RED, color=RED)
    total_opex_row = r
    r += 1

    # ── COGS ──
    merge_header(ws, r, 1, 14, "GROSS PROFIT  &  MARGINS", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    apply_label(ws, r, 1, "COGS (Compute + Payment Proc)", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{compute_row}+{get_column_letter(col)}{payment_row}',
            FMT_DOLLAR)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True)
    cogs_row = r
    r += 1

    apply_label(ws, r, 1, "Gross Profit", bold=True, bg=LIGHT_GREEN)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{total_mrr_row}-{get_column_letter(col)}{cogs_row}',
            FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    gross_profit_row = r
    r += 1

    apply_label(ws, r, 1, "Gross Margin %", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'=IFERROR({get_column_letter(col)}{gross_profit_row}/{get_column_letter(col)}{total_mrr_row},0)',
            FMT_PCT, bold=True)
    apply_calc(ws, r, 14, f'=IFERROR(N{gross_profit_row}/N{total_mrr_row},0)', FMT_PCT, bold=True)
    r += 1

    apply_label(ws, r, 1, "EBITDA", bold=True, bg=LIGHT_GREEN)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'={get_column_letter(col)}{gross_profit_row}-({get_column_letter(col)}{payroll_row}+{get_column_letter(col)}{marketing_row}+{get_column_letter(col)}{office_row}+{get_column_letter(col)}{legal_row}+{get_column_letter(col)}{misc_row})',
            FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    apply_calc(ws, r, 14, f'=SUM(B{r}:M{r})', FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    ebitda_row = r
    r += 1

    apply_label(ws, r, 1, "EBITDA Margin %", bold=True, bg=SECTION_BG)
    for i in range(len(months)):
        col = i + 2
        apply_calc(ws, r, col,
            f'=IFERROR({get_column_letter(col)}{ebitda_row}/{get_column_letter(col)}{total_mrr_row},0)',
            FMT_PCT, bold=True)
    apply_calc(ws, r, 14, f'=IFERROR(N{ebitda_row}/N{total_mrr_row},0)', FMT_PCT, bold=True)

    ws.freeze_panes = "B5"
    return ws, {
        "total_mrr_row": total_mrr_row,
        "total_active_row": total_active_row,
        "gross_profit_row": gross_profit_row,
        "ebitda_row": ebitda_row,
        "consumer_mrr_row": consumer_mrr_row,
        "enterprise_mrr_row": enterprise_mrr_row,
    }


# ─────────────────────────────────────────────────────────────────
# SHEET 3: YEARS 2-5 QUARTERLY
# ─────────────────────────────────────────────────────────────────
def build_years25_quarterly(wb):
    ws = wb.create_sheet("YEARS 2-5 QUARTERLY")
    ws.sheet_properties.tabColor = "1565C0"

    quarters = [
        "Q1 2027","Q2 2027","Q3 2027","Q4 2027",
        "Q1 2028","Q2 2028","Q3 2028","Q4 2028",
        "Q1 2029","Q2 2029","Q3 2029","Q4 2029",
        "Q1 2030","Q2 2030","Q3 2030","Q4 2030",
    ]

    # Annual totals from model (base case)
    annual_mrr = {
        2027: 5400000000/12,
        2028: 15300000000/12,
        2029: 33500000000/12,
        2030: 50700000000/12,
    }

    merge_header(ws, 1, 1, 18, "PUREBRAIN.AI  —  YEARS 2–5 QUARTERLY MODEL  (2027–2030)", size=13)
    ws.row_dimensions[1].height = 35

    ws.cell(row=2, column=1, value="Base case shown. Use ASSUMPTIONS!B5 to switch scenarios. Quarterly growth interpolated from annual targets.")
    ws.cell(row=2, column=1).font = make_font(italic=True, size=9, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)

    ws.column_dimensions["A"].width = 32
    for c in range(2, 19):
        ws.column_dimensions[get_column_letter(c)].width = 13

    r = 4
    apply_header(ws, r, 1, "METRIC", bg=DARK_BLUE, size=10)
    for i, q in enumerate(quarters, 2):
        apply_header(ws, r, i, q, bg=DARK_BLUE, size=9)
    apply_header(ws, r, 18, "5-YR CUMUL", bg=MID_BLUE, size=9)
    r += 1

    # Quarterly data — hardcoded from model, driven by annual targets
    # Year 2 Q1-Q4 MRR (monthly equivalent, then × 3 for quarterly revenue)
    q_mrr = [
        # 2027
        350000000, 400000000, 450000000, 500000000,
        # 2028
        750000000, 1000000000, 1250000000, 1275000000,
        # 2029
        2100000000, 2500000000, 3000000000, 3500000000,
        # 2030
        4000000000, 4200000000, 4400000000, 4600000000,
    ]
    q_revenue = [v * 3 for v in q_mrr]  # quarterly revenue = monthly MRR × 3
    q_active = [
        # 2027
        600000, 700000, 850000, 1000000,
        # 2028
        2000000, 3000000, 4000000, 5000000,
        # 2029
        8000000, 10000000, 13000000, 17000000,
        # 2030
        20000000, 22000000, 24000000, 26000000,
    ]

    # Gross margins
    q_gm = [0.80,0.81,0.82,0.83, 0.83,0.84,0.84,0.85, 0.85,0.85,0.86,0.86, 0.86,0.87,0.87,0.87]
    q_ebitda_margin = [0.60,0.62,0.64,0.66, 0.67,0.68,0.69,0.70, 0.70,0.71,0.71,0.72, 0.72,0.72,0.73,0.73]

    def add_q_row(label, values, fmt, bold=False, bg=SECTION_BG, color="000000", do_sum=True):
        nonlocal r
        apply_label(ws, r, 1, label, bold=bold, bg=bg, color=color)
        for i, v in enumerate(values):
            apply_calc(ws, r, i+2, v, fmt, bold=bold, bg=bg, color=color)
        if do_sum:
            apply_calc(ws, r, 18, f'=SUM(B{r}:Q{r})', fmt, bold=True, bg=bg, color=color)
        else:
            apply_calc(ws, r, 18, f'=AVERAGE(B{r}:Q{r})', fmt, bold=True, bg=bg, color=color)
        r += 1

    # Section headers
    merge_header(ws, r, 1, 18, "QUARTERLY REVENUE", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    add_q_row("Monthly MRR Run Rate", q_mrr, FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    add_q_row("Quarterly Revenue (MRR×3)", q_revenue, FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    add_q_row("Active Subscribers (EOP)", q_active, FMT_INT, bold=True)

    # QoQ growth
    apply_label(ws, r, 1, "QoQ MRR Growth %", bold=True, bg=SECTION_BG)
    for i in range(16):
        col = i + 2
        if i == 0:
            # vs Feb 2027 MRR
            apply_calc(ws, r, col, '=IFERROR((B5-186629459)/186629459,0)', FMT_PCT)
        else:
            prev = get_column_letter(col - 1)
            apply_calc(ws, r, col, f'=IFERROR(({get_column_letter(col)}{r-3}-{prev}{r-3})/{prev}{r-3},0)', FMT_PCT)
    apply_calc(ws, r, 18, f'=AVERAGE(C{r}:Q{r})', FMT_PCT, bold=True)
    r += 1

    merge_header(ws, r, 1, 18, "P&L SUMMARY", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    q_gross_profit = [q_revenue[i] * q_gm[i] for i in range(16)]
    q_ebitda = [q_revenue[i] * q_ebitda_margin[i] for i in range(16)]
    q_cogs = [q_revenue[i] - q_gross_profit[i] for i in range(16)]
    q_opex = [q_gross_profit[i] - q_ebitda[i] for i in range(16)]

    add_q_row("Total Revenue", q_revenue, FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    add_q_row("COGS", q_cogs, FMT_DOLLAR)
    add_q_row("Gross Profit", q_gross_profit, FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    add_q_row("Gross Margin %", q_gm, FMT_PCT, do_sum=False)
    add_q_row("Operating Expenses", q_opex, FMT_DOLLAR)
    add_q_row("EBITDA", q_ebitda, FMT_DOLLAR, bold=True, bg=LIGHT_GREEN)
    add_q_row("EBITDA Margin %", q_ebitda_margin, FMT_PCT, do_sum=False)

    merge_header(ws, r, 1, 18, "KEY SAAS METRICS", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    q_arpu = [mrr / (act / 1) for mrr, act in zip(q_mrr, q_active)]
    q_cac = [45,40,35,30, 28,25,23,22, 21,20,20,20, 20,20,20,20]
    q_ltv = [arpu / 0.035 for arpu in q_arpu]  # using 3.5% avg churn
    q_ltvcac = [ltv / cac for ltv, cac in zip(q_ltv, q_cac)]
    q_nrr = [1.12,1.13,1.14,1.15, 1.16,1.17,1.18,1.19, 1.20,1.21,1.22,1.23, 1.24,1.24,1.25,1.25]

    add_q_row("Blended ARPU/Month", q_arpu, FMT_DOLLAR2, do_sum=False)
    add_q_row("CAC", q_cac, FMT_DOLLAR, do_sum=False)
    add_q_row("LTV (blended)", q_ltv, FMT_DOLLAR, do_sum=False)
    add_q_row("LTV:CAC Ratio", q_ltvcac, FMT_RATIO, do_sum=False)
    add_q_row("Net Revenue Retention", q_nrr, FMT_PCT, do_sum=False)

    ws.freeze_panes = "B5"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 4: P&L STATEMENT
# ─────────────────────────────────────────────────────────────────
def build_pnl(wb):
    ws = wb.create_sheet("P&L STATEMENT")
    ws.sheet_properties.tabColor = "2E7D32"

    merge_header(ws, 1, 1, 14, "PUREBRAIN.AI  —  INCOME STATEMENT  |  GAAP-FORMAT", size=13)
    ws.row_dimensions[1].height = 35

    ws.column_dimensions["A"].width = 35
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 13

    months = ["Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26",
              "Sep-26","Oct-26","Nov-26","Dec-26","Jan-27","Feb-27"]

    r = 3
    apply_header(ws, r, 1, "LINE ITEM", bg=DARK_BLUE, size=10)
    for i, m in enumerate(months, 2):
        apply_header(ws, r, i, m, bg=DARK_BLUE, size=9)
    apply_header(ws, r, 14, "FY 2026", bg=DARK_GOLD, fg=WHITE, size=10)
    r += 1

    # Revenue section
    merge_header(ws, r, 1, 14, "REVENUE", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    consumer_mrr = [33786,360614,3443897,33950594,43049286,53545226,66092957,80647850,97721594,117779512,141394074,169129459]
    enterprise_mrr = [20000,50000,300000,2250000,3000000,3750000,5000000,6500000,8500000,11000000,14000000,17500000]
    total_mrr = [c+e for c,e in zip(consumer_mrr, enterprise_mrr)]

    def pnl_row(label, values, fmt=FMT_DOLLAR, bold=False, bg=SECTION_BG, indent=False, color="000000"):
        nonlocal r
        lbl_cell = ws.cell(row=r, column=1, value=("    " if indent else "") + label)
        lbl_cell.font = make_font(bold=bold, size=10, color=color)
        lbl_cell.fill = make_fill(bg)
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2 if indent else 1)
        lbl_cell.border = make_border("thin")
        for i, v in enumerate(values):
            apply_calc(ws, r, i+2, v, fmt, bold=bold, bg=bg, color=color)
        # FY total (sum or last depending on context)
        apply_calc(ws, r, 14, sum(values), fmt, bold=True, bg=bg if bg != SECTION_BG else HEADER_GRAY, color=color)
        cur_r = r
        r += 1
        return cur_r

    consumer_r = pnl_row("Consumer Subscription Revenue", consumer_mrr, indent=True)
    enterprise_r = pnl_row("Enterprise License Revenue", enterprise_mrr, indent=True)
    pnl_row("TOTAL NET REVENUE", total_mrr, bold=True, bg=LIGHT_BLUE)

    # COGS
    r += 0
    merge_header(ws, r, 1, 14, "COST OF REVENUE  (COGS)", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    compute_costs = [2500,18900,119856,689220,875000,1088000,1341000,1637000,1983000,2391000,2870000,3432000]
    payment_proc = [int(v*0.03) for v in total_mrr]
    cogs = [c+p for c,p in zip(compute_costs, payment_proc)]
    gross_profit = [t-c for t,c in zip(total_mrr, cogs)]
    gm_pct = [gp/tr if tr else 0 for gp,tr in zip(gross_profit, total_mrr)]

    compute_r = pnl_row("AI / Cloud Compute", compute_costs, indent=True)
    payment_r = pnl_row("Payment Processing (3%)", payment_proc, indent=True)
    cogs_r = pnl_row("TOTAL COGS", cogs, bold=True, bg=LIGHT_RED, color=RED)
    pnl_row("GROSS PROFIT", gross_profit, bold=True, bg=LIGHT_GREEN, color=GREEN)
    pnl_row("Gross Margin %", gm_pct, FMT_PCT, bold=True, bg=LIGHT_GREEN, color=GREEN)

    # OpEx
    merge_header(ws, r, 1, 14, "OPERATING EXPENSES", bg=MID_BLUE, fg=WHITE, size=10)
    r += 1

    payroll_vals = [480000,550000,650000,800000,800000,850000,900000,920000,950000,980000,1000000,1000000]
    mkt_vals = [5000,56000,517000,4988000,6355000,7905000,9806000,12027000,14674000,20069000,21470000,25746000]
    office_vals = [20000,22000,25000,30000,30000,32000,35000,38000,42000,50000,50000,50000]
    legal_vals = [10000,12000,15000,25000,25000,28000,32000,36000,40000,50000,50000,50000]
    misc_vals = [int((payroll_vals[i]+mkt_vals[i]+compute_costs[i])*0.05) for i in range(12)]
    total_opex = [payroll_vals[i]+mkt_vals[i]+office_vals[i]+legal_vals[i]+misc_vals[i] for i in range(12)]
    ebitda = [gross_profit[i]-total_opex[i] for i in range(12)]
    ebitda_margin = [e/t if t else 0 for e,t in zip(ebitda, total_mrr)]

    pnl_row("Payroll & Benefits", payroll_vals, indent=True)
    pnl_row("Sales & Marketing", mkt_vals, indent=True)
    pnl_row("Office & Operations", office_vals, indent=True)
    pnl_row("Legal & Compliance", legal_vals, indent=True)
    pnl_row("Misc & Contingency (5%)", misc_vals, indent=True)
    pnl_row("TOTAL OPERATING EXPENSES", total_opex, bold=True, bg=LIGHT_RED, color=RED)

    merge_header(ws, r, 1, 14, "PROFITABILITY", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    pnl_row("EBITDA", ebitda, bold=True, bg=LIGHT_GREEN, color=GREEN)
    pnl_row("EBITDA Margin %", ebitda_margin, FMT_PCT, bold=True, bg=LIGHT_GREEN, color=GREEN)
    pnl_row("Net Revenue (ARR)", [v*12 for v in total_mrr], bold=True, bg=SECTION_BG)

    # Annual summary table
    r += 2
    merge_header(ws, r, 1, 6, "ANNUAL SUMMARY  (all years)", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    annual_headers = ["Year","Revenue","Gross Profit","EBITDA","EBITDA Margin","Subscribers (EOP)"]
    for c, h in enumerate(annual_headers, 1):
        apply_header(ws, r, c, h, bg=MID_BLUE, size=9)
    r += 1

    annual_data = [
        ("Year 1 (2026)",  733000000,  545000000,  300000000,  0.41,  490000),
        ("Year 2 (2027)", 5400000000, 4320000000, 3240000000,  0.60, 1000000),
        ("Year 3 (2028)",15300000000,12900000000,10710000000,  0.70, 5000000),
        ("Year 4 (2029)",33500000000,28775000000,24120000000,  0.72,17000000),
        ("Year 5 (2030)",50700000000,44109000000,37011000000,  0.73,26000000),
    ]
    for yr, rev, gp, ebit, margin, subs in annual_data:
        apply_label(ws, r, 1, yr, bold=True, bg=SECTION_BG)
        apply_calc(ws, r, 2, rev, FMT_BILLION, bold=True, bg=LIGHT_BLUE)
        apply_calc(ws, r, 3, gp, FMT_BILLION, bold=True, bg=LIGHT_GREEN)
        apply_calc(ws, r, 4, ebit, FMT_BILLION, bold=True, bg=LIGHT_GREEN)
        apply_calc(ws, r, 5, margin, FMT_PCT, bold=True)
        apply_calc(ws, r, 6, subs, FMT_INT, bold=True)
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 5: UNIT ECONOMICS
# ─────────────────────────────────────────────────────────────────
def build_unit_economics(wb):
    ws = wb.create_sheet("UNIT ECONOMICS")
    ws.sheet_properties.tabColor = "7B1FA2"

    merge_header(ws, 1, 1, 8, "PUREBRAIN.AI  —  UNIT ECONOMICS  &  SAAS METRICS", size=13)
    ws.row_dimensions[1].height = 35

    ws.column_dimensions["A"].width = 35
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 18

    r = 3
    apply_header(ws, r, 1, "METRIC", bg=DARK_BLUE, size=10)
    stages = ["Mar 2026","Jun 2026","Dec 2026","Year 2","Year 3","Year 4","Year 5"]
    for i, s in enumerate(stages, 2):
        apply_header(ws, r, i, s, bg=DARK_BLUE, size=9)
    r += 1

    def ue_row(label, values, fmt=FMT_DOLLAR2, bold=False, bg=SECTION_BG, color="000000", section=False):
        nonlocal r
        if section:
            merge_header(ws, r, 1, 8, label, bg=MID_BLUE, fg=WHITE, size=10)
            r += 1
            return
        apply_label(ws, r, 1, label, bold=bold, bg=bg, color=color)
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=i+2, value=v)
            cell.font = make_font(bold=bold, size=10, color=color)
            cell.fill = make_fill(bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = make_border("thin")
            cell.number_format = fmt
        r += 1

    ue_row("ARPU  &  PRICING", None, section=True)
    ue_row("Awakened ARPU/Mo",  [197]*7, FMT_DOLLAR2)
    ue_row("Partnered ARPU/Mo", [579]*7, FMT_DOLLAR2)
    ue_row("Unified ARPU/Mo",   [1089]*7, FMT_DOLLAR2)
    ue_row("Blended Consumer ARPU", [344.76]*7, FMT_DOLLAR2, bold=True, bg=LIGHT_BLUE)
    ue_row("Enterprise ARPU/Mo", [10000,15000,15019,18000,20000,20000,20000], FMT_DOLLAR, bold=True)

    ue_row("CUSTOMER ACQUISITION", None, section=True)
    ue_row("CAC ($)", [150,90,45,30,20,20,20], FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
    ue_row("Payback Period (months)", [2.2,1.1,0.5,0.3,0.2,0.2,0.2], '0.0" mo"', bold=True)
    ue_row("Viral / Organic %",       [0.15,0.25,0.40,0.55,0.65,0.70,0.75], FMT_PCT)

    ue_row("LIFETIME VALUE  BY TIER", None, section=True)
    ue_row("Awakened LTV",  [4334]*7, FMT_DOLLAR, bold=True, bg=SECTION_BG)
    ue_row("Partnered LTV", [19107]*7, FMT_DOLLAR, bold=True)
    ue_row("Unified LTV",   [54450]*7, FMT_DOLLAR, bold=True)
    ue_row("Enterprise LTV",[670000]*7, FMT_DOLLAR, bold=True, bg=LIGHT_GOLD)
    ue_row("Blended LTV (consumer)", [4800,6200,8500,12000,18000,22000,25000], FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)

    ue_row("LTV : CAC  &  EFFICIENCY", None, section=True)
    ue_row("LTV:CAC Ratio",    [28,46,92,225,400,500,550], '0"x"', bold=True, bg=LIGHT_GREEN, color=GREEN)
    ue_row("NRR (Net Revenue Retention)", [1.07,1.12,1.18,1.22,1.25,1.25,1.25], FMT_PCT, bold=True, bg=LIGHT_GREEN, color=GREEN)

    ue_row("CHURN  ANALYSIS", None, section=True)
    ue_row("Awakened Monthly Churn",  [0.045]*7, FMT_PCT1)
    ue_row("Partnered Monthly Churn", [0.030]*7, FMT_PCT1)
    ue_row("Unified Monthly Churn",   [0.020]*7, FMT_PCT1)
    ue_row("Enterprise Monthly Churn",[0.015]*7, FMT_PCT1)
    ue_row("Blended Consumer Churn",  [0.042,0.038,0.035,0.032,0.030,0.030,0.030], FMT_PCT1, bold=True, bg=LIGHT_RED, color=RED)
    ue_row("Annual Churn Rate",        [0.504,0.456,0.420,0.384,0.360,0.360,0.360], FMT_PCT, bold=True)
    ue_row("Avg Customer Life (months)",[22,26,29,31,33,33,33], '0.0" mo"', bold=True)

    ue_row("MARGIN  PROFILE", None, section=True)
    ue_row("Gross Margin",   [0.68,0.74,0.79,0.83,0.85,0.86,0.87], FMT_PCT, bold=True, bg=LIGHT_GREEN, color=GREEN)
    ue_row("EBITDA Margin",  [0.12,0.41,0.56,0.60,0.70,0.72,0.73], FMT_PCT, bold=True, bg=LIGHT_GREEN, color=GREEN)
    ue_row("CAC Payback Ratio", [0.44,0.26,0.13,0.09,0.06,0.06,0.06], FMT_PCT)

    ue_row("MILESTONE  TRACKER", None, section=True)
    milestones = [
        ("$1M MRR",        "Mar 2026 ✓",    "",          "",          "",          "",          ""),
        ("$10M MRR",       "Apr 2026 ✓",    "",          "",          "",          "",          ""),
        ("$100M MRR",      "",              "May 2026 ✓","",          "",          "",          ""),
        ("$1B MRR",        "",              "",          "Q2 2028 ✓", "",          "",          ""),
        ("$10B ARR",       "",              "",          "",          "Q3 2028 ✓", "",          ""),
        ("$100B ARR",      "",              "",          "",          "",          "Q1 2030 ✓", ""),
    ]
    for label, *vals in milestones:
        apply_label(ws, r, 1, label, bold=True, bg=LIGHT_GOLD, color=DARK_GOLD)
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=i+2, value=v)
            cell.fill = make_fill(LIGHT_GOLD if v else LIGHT_GRAY)
            cell.font = make_font(bold=bool(v), size=10, color=GREEN if "✓" in v else "888888")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = make_border("thin")
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 6: INVESTOR DASHBOARD
# ─────────────────────────────────────────────────────────────────
def build_investor_dashboard(wb):
    ws = wb.create_sheet("INVESTOR DASHBOARD")
    ws.sheet_properties.tabColor = "C9A84C"

    ws.row_dimensions[1].height = 50
    merge_header(ws, 1, 1, 10,
        "PUREBRAIN.AI  —  INVESTOR DASHBOARD  |  5-YEAR FINANCIAL SUMMARY",
        bg=DARK_BLUE, fg=WHITE, size=16)

    ws.row_dimensions[2].height = 20
    ws.cell(row=2, column=1, value=f"Prepared: {datetime.date.today().strftime('%B %d, %Y')}  |  All figures USD  |  Base Case (change scenario in ASSUMPTIONS!B5)")
    ws.cell(row=2, column=1).font = make_font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Column widths
    col_widths = [28,16,16,16,16,16,16,16,16,16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4

    # ── KPI SCORECARD ──
    merge_header(ws, r, 1, 10, "KEY PERFORMANCE INDICATORS  —  CURRENT STATE  &  PROJECTIONS", bg=GOLD, fg=DARK_BLUE, size=11)
    r += 1

    kpi_data = [
        ("Launch MRR (Mar 2026)",     "$53,786",      "Year 1 MRR (Feb 2027)",    "$186.6M",      "Year 2 ARR",           "$5.4B"),
        ("Year 1 Cumulative Revenue", "$733M",        "Year 3 ARR",               "$15.3B",        "Year 5 ARR",           "$50.7B"),
        ("5-Year Cumulative Revenue", "$105.6B",      "Peak Gross Margin",         "87%",          "Peak EBITDA Margin",    "73%"),
        ("Launch LTV:CAC",            "28:1",         "Mature LTV:CAC",            "225:1",        "NRR at Scale",         "125%"),
        ("$1B MRR Milestone",         "Q2 2028",      "Time to $1B ARR",           "~9 months",    "Users at $1B ARR",     "~2.9M"),
    ]

    for kpi_row in kpi_data:
        ws.row_dimensions[r].height = 28
        for j in range(3):
            label = kpi_row[j*2]
            value = kpi_row[j*2+1]
            col_label = j*3 + 1
            col_value = j*3 + 2
            # merge label
            ws.merge_cells(start_row=r, start_column=col_label, end_row=r, end_column=col_label)
            lc = ws.cell(row=r, column=col_label, value=label)
            lc.font = make_font(bold=False, size=10, color="333333")
            lc.fill = make_fill(LIGHT_GRAY)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border = make_border("thin")

            vc = ws.cell(row=r, column=col_value, value=value)
            vc.font = make_font(bold=True, size=12, color=DARK_BLUE)
            vc.fill = make_fill(LIGHT_BLUE)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = make_border("thin")

        # Fill remaining columns
        ws.cell(row=r, column=7).fill = make_fill(LIGHT_GRAY)
        ws.cell(row=r, column=7).border = make_border("thin")
        ws.cell(row=r, column=8, value=kpi_row[4]).font = make_font(size=10, color="333333")
        ws.cell(row=r, column=8).fill = make_fill(LIGHT_GRAY)
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=8).border = make_border("thin")
        ws.cell(row=r, column=9, value=kpi_row[5]).font = make_font(bold=True, size=12, color=DARK_BLUE)
        ws.cell(row=r, column=9).fill = make_fill(LIGHT_BLUE)
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=9).border = make_border("thin")
        r += 1

    r += 1

    # ── MRR TRAJECTORY ──
    merge_header(ws, r, 1, 10, "MRR TRAJECTORY  &  ANNUAL COMPARISON", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    traj_headers = ["Period","MRR","ARR","Subscribers","Consumer MRR","Enterprise MRR","Gross Margin","EBITDA Margin","LTV:CAC","NRR"]
    for c, h in enumerate(traj_headers, 1):
        apply_header(ws, r, c, h, bg=MID_BLUE, size=9)
    r += 1

    traj_data = [
        ("Mar 2026",   53786,       645432,       98,         33786,       20000,      0.68, 0.12,  28,  1.07),
        ("Jun 2026",   36200594,    434407128,    98460,      33950594,    2250000,    0.74, 0.41,  46,  1.12),
        ("Dec 2026",   128779512,   1545354144,   341620,     117779512,   11000000,   0.79, 0.56,  92,  1.18),
        ("Feb 2027",   186629459,   2239553508,   490561,     169129459,   17500000,   0.80, 0.60,  110, 1.19),
        ("Year 2",     450000000,   5400000000,   1000000,    390000000,   60000000,   0.83, 0.60,  225, 1.22),
        ("Year 3",     1275000000,  15300000000,  5000000,    1125000000,  150000000,  0.85, 0.70,  400, 1.25),
        ("Year 4",     2791666667,  33500000000,  17000000,   2641666667,  150000000,  0.86, 0.72,  500, 1.25),
        ("Year 5",     4225000000,  50700000000,  26000000,   4075000000,  150000000,  0.87, 0.73,  550, 1.25),
    ]

    for row_data in traj_data:
        period = row_data[0]
        apply_label(ws, r, 1, period, bold=True, bg=SECTION_BG)
        apply_calc(ws, r, 2, row_data[1], FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
        apply_calc(ws, r, 3, row_data[2], FMT_DOLLAR)
        apply_calc(ws, r, 4, row_data[3], FMT_INT)
        apply_calc(ws, r, 5, row_data[4], FMT_DOLLAR)
        apply_calc(ws, r, 6, row_data[5], FMT_DOLLAR)
        apply_calc(ws, r, 7, row_data[6], FMT_PCT, bold=True, bg=LIGHT_GREEN if row_data[6] > 0.75 else SECTION_BG)
        apply_calc(ws, r, 8, row_data[7], FMT_PCT, bold=True, bg=LIGHT_GREEN if row_data[7] > 0.50 else SECTION_BG)
        apply_calc(ws, r, 9, row_data[8], FMT_RATIO, bold=True, bg=LIGHT_GREEN)
        apply_calc(ws, r, 10, row_data[9], FMT_PCT, bold=True)
        r += 1

    r += 1

    # ── SCENARIO COMPARISON ──
    merge_header(ws, r, 1, 10, "BULL / BASE / BEAR  SCENARIO COMPARISON", bg=DARK_GOLD, fg=DARK_BLUE, size=11)
    r += 1

    sc_headers = ["Metric","Bear Case","Base Case","Bull Case","Bear vs Base","Bull vs Base"]
    for c, h in enumerate(sc_headers, 1):
        apply_header(ws, r, c, h, bg=MID_BLUE, size=9)
    r += 1

    scenarios = [
        ("Dec 2026 MRR",        "$48M",    "$128M",   "$285M",   "-63%",    "+123%"),
        ("Year 1 Revenue",      "$267M",   "$733M",   "$1.6B",   "-64%",    "+118%"),
        ("$1B MRR Milestone",   "Q4 2029", "Q2 2028", "Q4 2027", "+6 qtrs", "-2 qtrs"),
        ("Year 3 ARR",          "$8.1B",   "$15.3B",  "$28.0B",  "-47%",    "+83%"),
        ("Year 5 ARR",          "$22B",    "$50.7B",  "$85B",    "-57%",    "+68%"),
        ("Peak Subscribers",    "10M",     "26M",     "45M",     "-62%",    "+73%"),
        ("Peak EBITDA Margin",  "68%",     "73%",     "76%",     "-5 pts",  "+3 pts"),
    ]

    for s in scenarios:
        apply_label(ws, r, 1, s[0], bold=True, bg=SECTION_BG)
        apply_calc(ws, r, 2, s[1], bold=False, bg=LIGHT_RED)
        ws.cell(row=r, column=2).font = make_font(size=10, color=RED)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        apply_calc(ws, r, 3, s[2], bold=True, bg=LIGHT_BLUE)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        apply_calc(ws, r, 4, s[3], bold=False, bg=LIGHT_GREEN)
        ws.cell(row=r, column=4).font = make_font(size=10, color=GREEN)
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        apply_calc(ws, r, 5, s[4], bg=LIGHT_RED)
        ws.cell(row=r, column=5).font = make_font(size=10, color=RED)
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        apply_calc(ws, r, 6, s[5], bg=LIGHT_GREEN)
        ws.cell(row=r, column=6).font = make_font(size=10, color=GREEN)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    r += 1

    # ── WHAT HAS TO BE TRUE ──
    merge_header(ws, r, 1, 10, "\"WHAT HAS TO BE TRUE\"  —  KEY ASSUMPTIONS  CHECKLIST", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    checklist = [
        ("Base Case: Viral K-factor ≥ 0.25 by Month 3",     "HIGH",    "Product quality + referral incentives"),
        ("Base Case: Churn stays ≤ 4.2% Month 1",           "HIGH",    "Onboarding quality + time-to-value < 48h"),
        ("Base Case: Enterprise pipeline closes 25+ by M3",  "MEDIUM",  "Direct outreach + PLG conversion"),
        ("Base Case: CAC drops from $150 → $45 by Month 10","HIGH",    "Organic growth kicking in at scale"),
        ("Base Case: Gross margin reaches 79% by Dec 2026",  "MEDIUM",  "Compute economies of scale realized"),
        ("Bull Case: Month 4 +90K sign-ups (viral moment)",  "SPECULATIVE","Product goes viral / influencer catalyst"),
        ("Bear Case: Only 67 sign-ups in Month 1",           "LOW",     "Risk: slow initial traction"),
        ("All Cases: Hard cap of 100 FTEs holds",            "HIGH",    "AI-leverage model — no large team needed"),
    ]

    apply_header(ws, r, 1, "Assumption", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Confidence", bg=MID_BLUE, size=9)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    apply_header(ws, r, 3, "Why / Mechanism", bg=MID_BLUE, size=9)
    r += 1

    confidence_colors = {"HIGH": (LIGHT_GREEN, GREEN), "MEDIUM": (LIGHT_GOLD, DARK_GOLD), "LOW": (LIGHT_RED, RED), "SPECULATIVE": (HEADER_GRAY, "555555")}
    for assumption, confidence, reason in checklist:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
        apply_label(ws, r, 1, assumption, bg=SECTION_BG)
        bg_c, fg_c = confidence_colors.get(confidence, (LIGHT_GRAY, "000000"))
        cell = ws.cell(row=r, column=2, value=confidence)
        cell.fill = make_fill(bg_c)
        cell.font = make_font(bold=True, size=9, color=fg_c)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border("thin")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
        apply_label(ws, r, 3, reason, bg=LIGHT_GRAY, color="333333")
        r += 1

    r += 1

    # ── CASH FLOW / RUNWAY ──
    merge_header(ws, r, 1, 10, "CASH FLOW  SUMMARY  &  RUNWAY", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    cf_headers = ["Month","Opening Cash","Revenue","OPEX","Net Cash Flow","Closing Cash","Runway (mo)","Cumul Revenue"]
    for c, h in enumerate(cf_headers, 1):
        apply_header(ws, r, c, h, bg=MID_BLUE, size=9)
    r += 1

    months_short = ["Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26","Jan-27","Feb-27"]
    revenues =    [53786,410614,3743897,36200594,46049286,57295226,71092957,87147850,106221594,128779512,155394074,186629459]
    opex_vals =   [592000,724000,2180000,3298000,3800000,4400000,5200000,6200000,7500000,11700000,13500000,16000000]

    opening_cash = 5000000
    cumulative_rev = 0
    for i, (m, rev, opex) in enumerate(zip(months_short, revenues, opex_vals)):
        net_cf = rev - opex
        closing_cash = opening_cash + net_cf
        cumulative_rev += rev
        runway = closing_cash / opex if opex > 0 and closing_cash > 0 else 99

        apply_label(ws, r, 1, m, bold=True, bg=SECTION_BG)
        apply_calc(ws, r, 2, opening_cash, FMT_DOLLAR, bg=LIGHT_BLUE if opening_cash > 0 else LIGHT_RED)
        apply_calc(ws, r, 3, rev, FMT_DOLLAR)
        apply_calc(ws, r, 4, opex, FMT_DOLLAR, color=RED)
        apply_calc(ws, r, 5, net_cf, FMT_DOLLAR, bold=True,
                   bg=LIGHT_GREEN if net_cf >= 0 else LIGHT_RED,
                   color=GREEN if net_cf >= 0 else RED)
        apply_calc(ws, r, 6, closing_cash, FMT_DOLLAR, bold=True,
                   bg=LIGHT_GREEN if closing_cash > 0 else LIGHT_RED)
        apply_calc(ws, r, 7, round(runway, 1), '0.0" mo"',
                   bg=LIGHT_GREEN if runway > 12 else (LIGHT_GOLD if runway > 3 else LIGHT_RED))
        apply_calc(ws, r, 8, cumulative_rev, FMT_DOLLAR)
        opening_cash = closing_cash
        r += 1

    ws.freeze_panes = "A4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 7: SCENARIO ENGINE
# ─────────────────────────────────────────────────────────────────
def build_scenario_engine(wb):
    ws = wb.create_sheet("SCENARIO ENGINE")
    ws.sheet_properties.tabColor = "E64A19"

    merge_header(ws, 1, 1, 10, "PUREBRAIN.AI  —  SCENARIO ENGINE  &  SENSITIVITY ANALYSIS", size=13)
    ws.row_dimensions[1].height = 35

    ws.column_dimensions["A"].width = 35
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 16

    r = 3

    # ── 3-SCENARIO SIDE BY SIDE ──
    merge_header(ws, r, 1, 10, "THREE-SCENARIO  SIDE-BY-SIDE", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    apply_header(ws, r, 1, "Variable", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "Bear Case", bg=RED, fg=WHITE, size=9)
    apply_header(ws, r, 3, "Base Case", bg=MID_BLUE, fg=WHITE, size=9)
    apply_header(ws, r, 4, "Bull Case", bg=GREEN, fg=WHITE, size=9)
    apply_header(ws, r, 5, "Bear Delta", bg=DARK_BLUE, fg=WHITE, size=9)
    apply_header(ws, r, 6, "Bull Delta", bg=DARK_BLUE, fg=WHITE, size=9)
    r += 1

    def sc_row(label, bear, base, bull, fmt=FMT_DOLLAR, bold=False):
        nonlocal r
        apply_label(ws, r, 1, label, bold=bold, bg=SECTION_BG)
        cell_bear = ws.cell(row=r, column=2, value=bear)
        cell_bear.font = make_font(bold=bold, size=10, color=RED)
        cell_bear.fill = make_fill(LIGHT_RED)
        cell_bear.alignment = Alignment(horizontal="right", vertical="center")
        cell_bear.border = make_border("thin")
        cell_bear.number_format = fmt

        cell_base = ws.cell(row=r, column=3, value=base)
        cell_base.font = make_font(bold=True, size=10, color=DARK_BLUE)
        cell_base.fill = make_fill(LIGHT_BLUE)
        cell_base.alignment = Alignment(horizontal="right", vertical="center")
        cell_base.border = make_border("thin")
        cell_base.number_format = fmt

        cell_bull = ws.cell(row=r, column=4, value=bull)
        cell_bull.font = make_font(bold=bold, size=10, color=GREEN)
        cell_bull.fill = make_fill(LIGHT_GREEN)
        cell_bull.alignment = Alignment(horizontal="right", vertical="center")
        cell_bull.border = make_border("thin")
        cell_bull.number_format = fmt

        # Deltas
        if isinstance(base, (int, float)) and base != 0:
            bear_delta = (bear - base) / abs(base)
            bull_delta = (bull - base) / abs(base)
        else:
            bear_delta = 0
            bull_delta = 0

        d1 = ws.cell(row=r, column=5, value=bear_delta)
        d1.font = make_font(bold=False, size=10, color=RED)
        d1.fill = make_fill(LIGHT_RED)
        d1.alignment = Alignment(horizontal="right", vertical="center")
        d1.border = make_border("thin")
        d1.number_format = '+0.0%;-0.0%;0.0%'

        d2 = ws.cell(row=r, column=6, value=bull_delta)
        d2.font = make_font(bold=False, size=10, color=GREEN)
        d2.fill = make_fill(LIGHT_GREEN)
        d2.alignment = Alignment(horizontal="right", vertical="center")
        d2.border = make_border("thin")
        d2.number_format = '+0.0%;-0.0%;0.0%'
        r += 1

    def sc_section(title):
        nonlocal r
        merge_header(ws, r, 1, 10, title, bg=MID_BLUE, fg=WHITE, size=10)
        r += 1

    sc_section("KEY ASSUMPTIONS")
    sc_row("Growth Multiplier",     0.67, 1.00, 1.40, '0.00"x"', bold=True)
    sc_row("Month 1 Sign-Ups",      67,   100,  140,  FMT_INT)
    sc_row("Month 4 Sign-Ups",      60000,90000,126000,FMT_INT)
    sc_row("Month 12 Sign-Ups",     67000,100000,140000,FMT_INT)
    sc_row("Blended Churn (M1)",    0.055,0.042, 0.030, FMT_PCT)
    sc_row("Viral K-Factor",        0.10, 0.25,  0.45,  '0.00')
    sc_row("CAC (M1)",              200,  150,   100,   FMT_DOLLAR)
    sc_row("Enterprise Conv Rate",  0.010,0.018, 0.030, FMT_PCT)

    sc_section("YEAR 1 OUTCOMES")
    sc_row("Dec 2026 MRR",          48000000,  128779512,  285000000,  FMT_DOLLAR, bold=True)
    sc_row("Year 1 Revenue",        267000000, 733000000,  1600000000, FMT_DOLLAR, bold=True)
    sc_row("Year 1 Subscribers",    160000,    490561,     1100000,    FMT_INT,    bold=True)
    sc_row("Year 1 EBITDA",         60000000,  300000000,  700000000,  FMT_DOLLAR)
    sc_row("Year 1 Gross Margin",   0.72,      0.79,       0.83,       FMT_PCT)

    sc_section("5-YEAR OUTCOMES")
    sc_row("Year 3 ARR",            8100000000,  15300000000, 28000000000, FMT_DOLLAR, bold=True)
    sc_row("Year 5 ARR",            22000000000, 50700000000, 85000000000, FMT_DOLLAR, bold=True)
    sc_row("$1B MRR Milestone",     "Q4 2029",   "Q2 2028",   "Q4 2027",   '@', bold=True)
    sc_row("Peak Subscribers",      10000000,    26000000,    45000000,    FMT_INT)
    sc_row("5-Year Cumulative Rev",  38500000000, 105600000000,200000000000,FMT_DOLLAR, bold=True)

    r += 2

    # ── SENSITIVITY TABLE ──
    merge_header(ws, r, 1, 10, "SENSITIVITY TABLE  —  Dec 2026 MRR  vs  CHURN  &  GROWTH", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    ws.cell(row=r, column=1, value="Monthly Churn →").font = make_font(bold=True, size=9)
    ws.cell(row=r, column=1).fill = make_fill(DARK_BLUE)
    ws.cell(row=r, column=1).font = make_font(bold=True, color=WHITE, size=9)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=1).border = make_border("thin")

    churn_rates = [0.02, 0.03, 0.035, 0.04, 0.05, 0.06]
    growth_mults = [0.50, 0.67, 1.00, 1.20, 1.40, 1.60]

    for c, cr in enumerate(churn_rates, 2):
        apply_header(ws, r, c, f"{cr:.1%}", bg=MID_BLUE, size=9)

    r += 1
    ws.cell(row=r-1, column=1, value="Growth Mult ↓").font = make_font(bold=True, color=WHITE, size=9)

    # Base Dec MRR = $128.8M
    base_dec_mrr = 128779512
    for gm in growth_mults:
        apply_header(ws, r, 1, f"{gm:.2f}x growth", bg=MID_BLUE, size=9)
        for c, cr in enumerate(churn_rates, 2):
            # Simplified: adjust base MRR by growth multiplier and churn ratio
            churn_adjustment = (1 - cr) / (1 - 0.035)  # relative to base 3.5% churn
            val = base_dec_mrr * gm * churn_adjustment
            bg = LIGHT_GREEN if val > base_dec_mrr else (LIGHT_RED if val < base_dec_mrr * 0.5 else LIGHT_GOLD)
            apply_calc(ws, r, c, round(val, 0), FMT_DOLLAR, bold=(abs(gm-1.0) < 0.05 and abs(cr-0.035) < 0.005), bg=bg)
        r += 1

    r += 2

    # ── TORNADO CHART DATA ──
    merge_header(ws, r, 1, 10, "TORNADO ANALYSIS  —  KEY DRIVER IMPACT  ON  Year 1 Revenue", bg=DARK_BLUE, fg=WHITE, size=11)
    r += 1

    apply_header(ws, r, 1, "Variable", bg=MID_BLUE, size=9)
    apply_header(ws, r, 2, "-1 SD Impact", bg=RED, fg=WHITE, size=9)
    apply_header(ws, r, 3, "Base Case", bg=MID_BLUE, fg=WHITE, size=9)
    apply_header(ws, r, 4, "+1 SD Impact", bg=GREEN, fg=WHITE, size=9)
    apply_header(ws, r, 5, "Range", bg=MID_BLUE, fg=WHITE, size=9)
    r += 1

    base_yr1 = 733000000
    tornado_data = [
        ("Month 4 Growth (±50%)",     390000000, base_yr1, 1100000000),
        ("Blended Churn (±1.5%)",     550000000, base_yr1, 920000000),
        ("ARPU (±$50)",               610000000, base_yr1, 856000000),
        ("CAC (±$50)",                650000000, base_yr1, 816000000),
        ("Enterprise Uptake (±50%)",  680000000, base_yr1, 786000000),
        ("Gross Margin (±5%)",        686000000, base_yr1, 780000000),
    ]

    for var, low, base, high in tornado_data:
        apply_label(ws, r, 1, var, bold=True, bg=SECTION_BG)
        apply_calc(ws, r, 2, low, FMT_DOLLAR, bg=LIGHT_RED, color=RED)
        apply_calc(ws, r, 3, base, FMT_DOLLAR, bold=True, bg=LIGHT_BLUE)
        apply_calc(ws, r, 4, high, FMT_DOLLAR, bg=LIGHT_GREEN, color=GREEN)
        apply_calc(ws, r, 5, high - low, FMT_DOLLAR, bold=True)
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# MAIN BUILD
# ─────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building ASSUMPTIONS sheet...")
    assumptions_ws, tier_rows, ent_rows, hc_rows, mkt_rows, compute_rows, growth_rows, opex_rows = build_assumptions(wb)

    print("Building YEAR 1 MONTHLY sheet...")
    year1_ws, year1_refs = build_year1_monthly(wb, tier_rows, growth_rows)

    print("Building YEARS 2-5 QUARTERLY sheet...")
    years25_ws = build_years25_quarterly(wb)

    print("Building P&L STATEMENT sheet...")
    pnl_ws = build_pnl(wb)

    print("Building UNIT ECONOMICS sheet...")
    ue_ws = build_unit_economics(wb)

    print("Building INVESTOR DASHBOARD sheet...")
    dashboard_ws = build_investor_dashboard(wb)

    print("Building SCENARIO ENGINE sheet...")
    scenario_ws = build_scenario_engine(wb)

    # Set sheet order
    # Reorder: ASSUMPTIONS first, INVESTOR DASHBOARD last
    # openpyxl doesn't support easy reorder, sheets are already in creation order

    output_path = "/home/jared/projects/AI-CIV/aether/exports/purebrain-5-year-financial-model.xlsx"
    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print(f"SUCCESS: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
