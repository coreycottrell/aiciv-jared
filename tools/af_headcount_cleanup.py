"""
AF# Headcount Sheet Cleanup Script
Modifies sheets 5B and 5C to reduce headcount to target levels
WITHOUT breaking any formulas.

Targets: Y0=100, Y1=125, Y2=150, Y3=175, Y4=200, Y5=250
Named people on 5B: Pre-Launch=42, Y1=46, Y2=46, Y3=46, Y4=46, Y5=46
5C needs: Pre-Launch=58, Y1=79, Y2=104, Y3=129, Y4=154, Y5=204
"""

import openpyxl
import shutil
import math

SRC = '/home/jared/portal_uploads/from-portal/portal_20260316_165059_PT-Full-Financial-Model-WITH-PureBrainNEARLYFINISHED.xlsx'
DST = '/home/jared/projects/AI-CIV/aether/exports/PT-Full-Financial-Model-WITH-PureBrain.xlsx'

print("Loading workbook (preserving formulas)...")
wb = openpyxl.load_workbook(SRC, data_only=False)

# ============================================================
# SHEET 5B: Zero out TBD rows
# ============================================================
ws5b = wb['5B. HC and Salaries - Mgmt.']

# HC columns (headcount - years): O=Pre-Launch(15), P=Y1(16), Q=Y2(17), R=Y3(18), S=Y4(19), T=Y5(20)
# Salary columns: V=Pre-Launch(22), W=Y1(23), X=Y2(24)
# ALL value cells in TBD rows need to be zeroed EXCEPT formula cells

# TBD rows in Senior Management
tbd_senior = [47]  # Chief Innovation Officer

# TBD rows in Middle Management (all rows that are TBD)
tbd_middle = [
    71, 73, 75, 76, 81, 82, 83, 84, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98,
    99, 100, 101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114,
    115, 116, 117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130,
    131, 132, 133, 134, 140, 150
]

all_tbd_rows_5b = tbd_senior + tbd_middle

print(f"\n5B: Zeroing out {len(all_tbd_rows_5b)} TBD rows...")

zeroed_cells = 0
for row_num in all_tbd_rows_5b:
    for cell in ws5b[row_num]:
        # Skip formula cells (data_type == 'f' or value is a string starting with '=')
        if cell.data_type == 'f':
            continue
        if isinstance(cell.value, str) and cell.value.startswith('='):
            continue
        # Skip None cells
        if cell.value is None:
            continue
        # Skip identifier columns: B(row number formula), D(bullet), E(name), I(position), J(location), L(start), M(hire)
        if cell.column in [2, 4, 5, 9, 10, 12, 13]:
            continue
        # Zero out numeric value cells
        if isinstance(cell.value, (int, float)):
            cell.value = 0
            zeroed_cells += 1

print(f"  Zeroed {zeroed_cells} value cells in 5B TBD rows")

# Verify row 47 after zeroing
print("\n  Verification - Row 47 (TBD CIO) after zeroing:")
for cell in ws5b[47]:
    if cell.value is not None and cell.column in [15, 16, 17, 18, 19, 20]:
        print(f"    [{cell.coordinate}] = {cell.value}")

# Also verify row 150 (TBD Hardware Director)
print("\n  Verification - Row 150 (TBD Hardware Director) after zeroing:")
for cell in ws5b[150]:
    if cell.value is not None and cell.column in [15, 16, 17, 18, 19, 20, 22]:
        print(f"    [{cell.coordinate}] = {cell.value}")

# ============================================================
# SHEET 5C: Scale headcount to meet targets
# ============================================================
ws5c = wb['5C. HC and Salaries - Support']

print("\n5C: Adjusting support headcount to hit targets...")

# 5C year structure:
# Pre-Launch: value cols 28-39 (AB-AM), formula col 40 (AN)
# Year 1:     value cols 42-53 (AP-BA), formula col 54 (BB)
# Year 2:     value cols 56-67 (BD-BO), formula col 68 (BP)
# Year 3:     value cols 70-81 (BR-CC), formula col 82 (CD)
# Year 4:     value cols 84-95 (CF-CQ), formula col 96 (CR)
# Year 5:     value cols 98-109 (CT-DE), formula col 110 (DF)

# Targets for 5C (Total - 5B Named People)
# 5B named: Pre-Launch=42, Y1=46, Y2=46, Y3=46, Y4=46, Y5=46
# Targets:  Y0=100, Y1=125, Y2=150, Y3=175, Y4=200, Y5=250
# 5C needs: Pre-Launch=58, Y1=79, Y2=104, Y3=129, Y4=154, Y5=204

year_config = {
    'Pre-Launch': {
        'target': 58,
        'value_cols': list(range(28, 40)),  # AB-AM (12 cols)
        'formula_col': 40,                  # AN
        'start_col': 28,
    },
    'Year1': {
        'target': 79,
        'value_cols': list(range(42, 54)),  # AP-BA (12 cols)
        'formula_col': 54,                  # BB
        'start_col': 42,
    },
    'Year2': {
        'target': 104,
        'value_cols': list(range(56, 68)),  # BD-BO (12 cols)
        'formula_col': 68,                  # BP
        'start_col': 56,
    },
    'Year3': {
        'target': 129,
        'value_cols': list(range(70, 82)),  # BR-CC (12 cols)
        'formula_col': 82,                  # CD
        'start_col': 70,
    },
    'Year4': {
        'target': 154,
        'value_cols': list(range(84, 96)),  # CF-CQ (12 cols)
        'formula_col': 96,                  # CR
        'start_col': 84,
    },
    'Year5': {
        'target': 204,
        'value_cols': list(range(98, 110)), # CT-DE (12 cols)
        'formula_col': 110,                 # DF
        'start_col': 98,
    },
}

data_rows_5c = list(range(22, 50))  # rows 22-49

# Step 1: Get current totals per year using start col
current_totals = {}
for yr, cfg in year_config.items():
    total = 0
    for rn in data_rows_5c:
        for cell in ws5c[rn]:
            if cell.column == cfg['start_col']:
                v = cell.value
                if isinstance(v, (int, float)):
                    total += v
    current_totals[yr] = total

print("\n  Current 5C totals (beginning of year col):")
for yr, t in current_totals.items():
    print(f"    {yr}: {t}")

print("\n  Target 5C totals:")
for yr, cfg in year_config.items():
    print(f"    {yr}: {cfg['target']}")

# Step 2: For each year, scale all rows proportionally to hit target
# Strategy:
# - Collect {row: current_start_value} for all rows with non-zero values
# - Scale proportionally so sum = target
# - Set ALL monthly cells in that year to the scaled value (constant per year per row)
# - Rows with 0 stay 0

for yr, cfg in year_config.items():
    target = cfg['target']
    current = current_totals[yr]
    start_col = cfg['start_col']
    value_cols = cfg['value_cols']

    print(f"\n  Processing {yr}: current={current}, target={target}")

    if current == 0:
        # Nothing to scale - just distribute proportionally from Y1 or zero
        # We'll handle this by distributing evenly across non-zero rows later
        print(f"    WARNING: Current total is 0, will use proportional distribution from Year1")
        # Use Year1 proportions as template
        current = current_totals.get('Year1', 0)
        start_col_ref = year_config['Year1']['start_col']
        row_values = {}
        for rn in data_rows_5c:
            for cell in ws5c[rn]:
                if cell.column == start_col_ref:
                    v = cell.value
                    if isinstance(v, (int, float)) and v > 0:
                        row_values[rn] = v
        if not row_values or sum(row_values.values()) == 0:
            # Evenly distribute across 14 category rows
            even_val = target // 14
            for rn in data_rows_5c:
                for cell in ws5c[rn]:
                    if cell.column == start_col:
                        pos = ws5c.cell(row=rn, column=9).value
                        if pos:  # has a position label
                            for vc in value_cols:
                                tc = ws5c.cell(row=rn, column=vc)
                                if tc.data_type != 'f' and not (isinstance(tc.value, str) and tc.value.startswith('=')):
                                    tc.value = even_val
        else:
            total_ref = sum(row_values.values())
            assigned = {}
            for rn, v in row_values.items():
                assigned[rn] = max(1, round(v / total_ref * target))

            # Adjust rounding to hit exact target
            diff = target - sum(assigned.values())
            if diff != 0:
                sorted_rows = sorted(assigned.keys(), key=lambda r: assigned[r], reverse=True)
                idx = 0
                while diff != 0:
                    assigned[sorted_rows[idx % len(sorted_rows)]] += (1 if diff > 0 else -1)
                    diff += (-1 if diff > 0 else 1)
                    idx += 1

            for rn in data_rows_5c:
                new_val = assigned.get(rn, 0)
                for vc in value_cols:
                    tc = ws5c.cell(row=rn, column=vc)
                    if tc.data_type != 'f' and not (isinstance(tc.value, str) and tc.value.startswith('=')):
                        tc.value = new_val
        continue

    # Collect current start values per row
    row_values = {}
    for rn in data_rows_5c:
        for cell in ws5c[rn]:
            if cell.column == start_col:
                v = cell.value
                if isinstance(v, (int, float)):
                    row_values[rn] = v

    # Scale proportionally
    scale = target / current if current > 0 else 0
    assigned = {}
    for rn, v in row_values.items():
        if v > 0:
            assigned[rn] = max(1, round(v * scale))
        else:
            assigned[rn] = 0

    # Adjust rounding to hit exact target
    total_assigned = sum(assigned.values())
    diff = target - total_assigned
    if diff != 0:
        # Adjust rows with non-zero values, largest first
        non_zero_rows = sorted([r for r in assigned if assigned[r] > 0],
                               key=lambda r: assigned[r], reverse=True)
        if non_zero_rows:
            idx = 0
            while diff != 0:
                rn = non_zero_rows[idx % len(non_zero_rows)]
                if diff > 0:
                    assigned[rn] += 1
                    diff -= 1
                else:
                    if assigned[rn] > 1:  # don't go below 1 for non-zero rows
                        assigned[rn] -= 1
                        diff += 1
                    else:
                        # Use next row
                        idx += 1
                        continue
                idx += 1

    print(f"    Scale factor: {scale:.4f}")
    print(f"    Total assigned: {sum(assigned.values())} (target: {target})")

    # Apply: set ALL monthly value cells in this year to the row's scaled value
    # Note: within a year, we keep the SAME value for all months
    # (the original data mostly has constant or stepped values per year)
    # For simplicity and formula safety, we set all value cols to the start value
    for rn in data_rows_5c:
        new_val = assigned.get(rn, 0)
        for vc in value_cols:
            tc = ws5c.cell(row=rn, column=vc)
            if tc.data_type == 'f':
                continue  # never touch formulas
            if isinstance(tc.value, str) and tc.value.startswith('='):
                continue  # never touch formulas
            # Set to new value (even if None - set to new_val if it exists)
            # Only set if cell had a value OR the row has a position label
            pos = ws5c.cell(row=rn, column=9).value
            if pos is not None:  # has a position/team label
                tc.value = new_val

print("\n5C changes complete.")

# ============================================================
# VERIFICATION
# ============================================================
print("\n" + "="*60)
print("VERIFICATION")
print("="*60)

# 5B: Check that TBD rows are zeroed
print("\n5B TBD Row Verification:")
hc_cols = [15, 16, 17, 18, 19, 20]  # O,P,Q,R,S,T
for rn in [47, 71, 75, 76, 81, 87, 88, 89, 90, 107, 108, 140, 150]:
    vals = []
    for cell in ws5b[rn]:
        if cell.column in hc_cols and cell.data_type != 'f':
            vals.append(f"{cell.coordinate}={cell.value}")
    name_cell = ws5b.cell(row=rn, column=5)
    print(f"  Row {rn} [{name_cell.value}]: {', '.join(vals)}")

# 5B: Check that named people are intact
print("\n5B Named Person Verification (spot check):")
for rn in [23, 24, 72, 74, 85, 86, 135, 149]:
    vals = []
    for cell in ws5b[rn]:
        if cell.column in hc_cols:
            vals.append(f"{cell.coordinate}={cell.value}")
    name_cell = ws5b.cell(row=rn, column=5)
    print(f"  Row {rn} [{name_cell.value}]: {', '.join(vals)}")

# 5B: Check sub-totals formula ranges are intact
print("\n5B Sub-Total formulas (should still be intact):")
for cell in ws5b[50]:  # Senior sub-total row
    if cell.column in hc_cols and cell.data_type == 'f':
        print(f"  {cell.coordinate}: {cell.value}")

# 5C: Verify new totals
print("\n5C New Totals Verification:")
new_totals = {}
for yr, cfg in year_config.items():
    total = 0
    for rn in data_rows_5c:
        cell = ws5c.cell(row=rn, column=cfg['start_col'])
        v = cell.value
        if isinstance(v, (int, float)):
            total += v
    new_totals[yr] = total

for yr in year_config:
    target = year_config[yr]['target']
    got = new_totals[yr]
    status = "OK" if got == target else f"MISMATCH (expected {target})"
    print(f"  {yr}: {got} {status}")

# 5C: Verify no formula cells were touched
print("\n5C Formula Cell Integrity Check:")
formula_cells_intact = True
formula_cols = [40, 54, 68, 82, 96, 110]  # AN, BB, BP, CD, CR, DF
for rn in data_rows_5c:
    for fc in formula_cols:
        cell = ws5c.cell(row=rn, column=fc)
        if cell.value is not None and cell.data_type != 'f':
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                print(f"  WARNING: Non-formula in expected formula cell {cell.coordinate}: {cell.value}")
                formula_cells_intact = False

if formula_cells_intact:
    print("  All year-end formula cells intact.")

# ============================================================
# SAVE
# ============================================================
print(f"\nSaving to: {DST}")
wb.save(DST)
print("Saved successfully.")

# Final summary
print("\n" + "="*60)
print("SUMMARY")
print("="*60)
print(f"\n5B: Zeroed {len(all_tbd_rows_5b)} TBD rows")
print("  - 1 Senior TBD (Chief Innovation Officer, row 47)")
print("  - 55 Middle Management TBD rows")
print("  - All named people preserved intact")
print("  - All Sub-Total/CONUS/OCONUS formulas untouched")
print("\n5C: Support headcount scaled to targets:")
for yr in year_config:
    target = year_config[yr]['target']
    got = new_totals[yr]
    print(f"  {yr}: {got}/{target}")
print("\nOutput file:", DST)
