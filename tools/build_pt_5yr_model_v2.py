"""
Pure Technology Unified 5-Year Financial Model v2
Rebuilt with:
- Dynamic INPUTS sheet (all assumptions in one place)
- Separate P&L per business unit
- PureBrain 97% gross margin / 70-80%+ EBITDA
- PT Core 30-40% NET margin (Jared confirmed)
- Blended margins improving as PureBrain grows
- All cells formula-driven referencing Inputs
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import os

# ─────────────────────────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────────────────────────
DARK_BG     = "0A0A1A"
PT_BLUE     = "2A93C1"
PT_ORANGE   = "F1420B"
GOLD        = "D4AF37"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MID_GRAY    = "CCCCCC"
DARK_GRAY   = "333333"
GREEN       = "1A7A4A"
GREEN_LIGHT = "E8F5EE"
RED_LIGHT   = "FFF0EE"
BLUE_LIGHT  = "EFF8FF"
ORANGE_LIGHT= "FFF4F0"
GOLD_LIGHT  = "FFF9ED"
SECTION_HDR = "1A3A5C"  # dark navy for section headers
INPUT_FILL  = "FFFFF0"  # pale yellow for input cells
FORMULA_FILL= "F0F8FF"  # pale blue for formula cells

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="thin", color="999999")
    return Border(bottom=s)

def border_thick_bottom():
    s = Side(style="medium", color="555555")
    return Border(bottom=s)

def style_header(ws, row, col, value, bg=SECTION_HDR, fg=WHITE,
                 size=12, bold=True, span=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=size)
    cell.alignment = align("center")
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col+span-1
        )
    return cell

def style_subheader(ws, row, col, value, bg=PT_BLUE, fg=WHITE,
                    size=10, span=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=size)
    cell.alignment = align("center")
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col+span-1
        )
    return cell

def style_input(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(INPUT_FILL)
    cell.font = font(bold=False, color="1A3A5C", size=10)
    cell.alignment = align("right")
    cell.border = border_thin()
    if fmt:
        cell.number_format = fmt
    return cell

def style_label(ws, row, col, value, bold=False, indent=0):
    cell = ws.cell(row=row, column=col, value=(" " * indent * 2) + value)
    cell.font = font(bold=bold, color=DARK_GRAY, size=10)
    cell.alignment = align("left")
    return cell

def style_formula(ws, row, col, formula, fmt=None, bg=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = font(color="1A1A4A", size=10)
    cell.alignment = align("right")
    cell.border = border_thin()
    if bg:
        cell.fill = fill(bg)
    if fmt:
        cell.number_format = fmt
    return cell

def style_total(ws, row, col, formula, fmt=None, bg=GREEN, fg=WHITE):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=10)
    cell.alignment = align("right")
    cell.border = border_thin()
    if fmt:
        cell.number_format = fmt
    return cell

def style_pct_total(ws, row, col, formula, bg=GREEN, fg=WHITE):
    style_total(ws, row, col, formula, fmt="0.0%", bg=bg, fg=fg)

USD = '#,##0'
USD2 = '#,##0.00'
USDM = '#,##0,,"M"'
PCT = '0.0%'
NUM = '#,##0'
MULT = '0.0"x"'

# ─────────────────────────────────────────────────────────────────
# SHEET 1: INPUTS
# ─────────────────────────────────────────────────────────────────
def build_inputs(wb):
    ws = wb.create_sheet("INPUTS")
    ws.tab_color = PT_ORANGE

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28

    # ── TITLE ──
    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — UNIFIED 5-YEAR FINANCIAL MODEL v2 INPUTS",
        bg=DARK_BG, fg=PT_ORANGE, size=14, span=6)
    ws.row_dimensions[1].height = 30

    style_header(ws, 2, 1,
        "ALL ASSUMPTIONS LIVE HERE — Every other sheet references this page",
        bg=SECTION_HDR, fg=WHITE, size=10, span=6)

    # ── SECTION: PUREBRAIN PRICING ──
    r = 4
    style_header(ws, r, 1, "PUREBRAIN PRICING TIERS", bg=PT_BLUE, fg=WHITE,
                 size=11, span=3)
    style_header(ws, r, 4, "INTRO PRICING (First 2 Months)", bg=PT_BLUE, fg=WHITE,
                 size=11, span=3)

    r += 1
    for col, hdr in enumerate(["Tier", "Monthly Price", "Notes"], 1):
        style_subheader(ws, r, col, hdr, span=1)
    for col, hdr in enumerate(["Tier", "Intro Price/Mo", "Duration (Months)"], 4):
        style_subheader(ws, r, col, hdr, span=1)

    r += 1
    tiers = [
        ("Bonded",    197,   "Entry tier — individual"),
        ("Partnered", 579,   "Small team / advanced"),
        ("Unified",   1089,  "Mid-market / department"),
        ("Enterprise",5500,  "Large org — custom avg"),
    ]
    intro_prices = [149, 499, 999, 4000]
    intro_dur = [2, 2, 2, 2]

    for i, (tier, price, note) in enumerate(tiers):
        style_label(ws, r+i, 1, tier, bold=True)
        style_input(ws, r+i, 2, price, USD2)
        ws.cell(r+i, 3, note).font = font(color="555555", size=9, italic=True)
        style_label(ws, r+i, 4, tier, bold=True)
        style_input(ws, r+i, 5, intro_prices[i], USD2)
        style_input(ws, r+i, 6, intro_dur[i], NUM)

    # Named ranges (manual approach — use consistent row references)
    # Row offsets for pricing (used in formula sheets):
    # Bonded price = B7, Partnered = B8, Unified = B9, Enterprise = B10
    # Intro Bonded = E7, etc.

    r += 5

    # ── SECTION: TIER MIX & CHURN ──
    style_header(ws, r, 1, "TIER MIX & CHURN BY YEAR", bg=SECTION_HDR, fg=WHITE,
                 size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1 (2026)", "Y2 (2027)", "Y3 (2028)", "Y4 (2029)", "Y5 (2030)"], 1):
        style_subheader(ws, r, col, hdr)

    r += 1
    mix_data = [
        ("Bonded % of mix",     0.60, 0.55, 0.50, 0.45, 0.40),
        ("Partnered % of mix",  0.25, 0.25, 0.25, 0.25, 0.25),
        ("Unified % of mix",    0.12, 0.16, 0.20, 0.24, 0.28),
        ("Enterprise % of mix", 0.03, 0.04, 0.05, 0.06, 0.07),
        ("Monthly Churn Rate",  0.030,0.025,0.020,0.015,0.012),
        ("Net Revenue Retention",1.05, 1.08, 1.10, 1.12, 1.15),
        ("Intro Pricing Months",    2,    2,    2,    1,    1),
    ]
    mix_rows = {}
    for label, *vals in mix_data:
        style_label(ws, r, 1, label, bold=("%" in label or "Churn" in label or "NRR" in label.upper()))
        for c, v in enumerate(vals, 2):
            fmt = PCT if ("%" in label or "Churn" in label or "NRR" in label.upper() or "Retention" in label) else NUM
            style_input(ws, r, c, v, fmt)
        mix_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: PUREBRAIN CUSTOMER ACQUISITION ──
    style_header(ws, r, 1, "PUREBRAIN CUSTOMER ACQUISITION", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1 (2026)", "Y2 (2027)", "Y3 (2028)", "Y4 (2029)", "Y5 (2030)"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    acq_data = [
        ("Launch Customers (Month 1)",  500,   0,     0,     0,     0),
        ("Monthly New Customer Growth", 200,   400,   800,   1500,  2500),
        ("Avg Monthly New Customers",   200,   400,   800,   1500,  2500),
        ("Year-End Total Customers",    2900,  7700,  17300, 35300, 65300),
    ]
    acq_rows = {}
    for label, *vals in acq_data:
        style_label(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            style_input(ws, r, c, v, NUM)
        acq_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: PUREBRAIN COGS ──
    style_header(ws, r, 1, "PUREBRAIN COGS (Per Customer / Month)", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Cost Item", "Y1", "Y2", "Y3", "Y4", "Y5"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    cogs_data = [
        ("Claude API Cost / Customer / Mo",  8.00, 7.00, 6.00, 5.50, 5.00),
        ("Infrastructure / Customer / Mo",   4.00, 3.50, 3.00, 2.75, 2.50),
        ("Support / Customer / Mo",          2.00, 1.75, 1.50, 1.25, 1.00),
        ("Total COGS / Customer / Mo",      14.00,12.25,10.50, 9.50, 8.50),
    ]
    cogs_rows = {}
    for i, (label, *vals) in enumerate(cogs_data):
        is_total = "Total" in label
        style_label(ws, r, 1, label, bold=is_total)
        for c, v in enumerate(vals, 2):
            if is_total:
                # Formula: sum of rows above
                style_total(ws, r, c,
                    f"=INPUTS!{get_column_letter(c)}{r-3}+INPUTS!{get_column_letter(c)}{r-2}+INPUTS!{get_column_letter(c)}{r-1}",
                    fmt=USD2, bg="1A5C3A", fg=WHITE)
            else:
                style_input(ws, r, c, v, USD2)
        cogs_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: PT CORE ──
    style_header(ws, r, 1, "PT CORE BUSINESS ASSUMPTIONS", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1 (2026)", "Y2 (2027)", "Y3 (2028)", "Y4 (2029)", "Y5 (2030)"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    pt_data = [
        ("Revenue (Full Year Run Rate $M)",462.0, 508.0, 559.0, 615.0, 677.0),
        ("PT Core Gross Margin %",         0.55,  0.56,  0.57,  0.58,  0.59),
        ("PT Core NET Margin % (Jared)",   0.35,  0.36,  0.37,  0.38,  0.39),
        ("PT Core OpEx % of Revenue",      0.20,  0.20,  0.20,  0.20,  0.20),
        ("PT Core COGS % of Revenue",      0.45,  0.44,  0.43,  0.42,  0.41),
        ("Ramp Start Month (1=Jan 2026)",    6,     1,     1,     1,     1),
        ("Ramp Factor Y1 (partial year)",  0.583,  1.0,   1.0,   1.0,   1.0),
    ]
    pt_rows = {}
    for label, *vals in pt_data:
        style_label(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            fmt = PCT if ("%" in label) else (NUM if ("Month" in label or "Factor" in label or isinstance(v, int)) else USD2)
            if "Revenue" in label:
                fmt = '#,##0.0"M"'
            style_input(ws, r, c, v, fmt)
        pt_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: PUREMARKETING ──
    style_header(ws, r, 1, "PUREMARKETING.AI ASSUMPTIONS", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1", "Y2", "Y3", "Y4", "Y5"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    pm_data = [
        ("Revenue ($M)",        1.8,  2.5,  3.5,  4.2,  5.0),
        ("Gross Margin %",      0.45, 0.46, 0.47, 0.48, 0.50),
        ("OpEx % of Revenue",   0.30, 0.28, 0.26, 0.24, 0.22),
    ]
    pm_rows = {}
    for label, *vals in pm_data:
        style_label(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            fmt = PCT if "%" in label else '#,##0.0"M"'
            style_input(ws, r, c, v, fmt)
        pm_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: PURE INFLUENCE ──
    style_header(ws, r, 1, "PURE INFLUENCE PLATFORM", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1", "Y2", "Y3", "Y4", "Y5"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    pi_data = [
        ("Revenue ($M)",        0.0,  0.3,  1.2,  3.0,  6.5),
        ("Gross Margin %",      0.0,  0.70, 0.73, 0.75, 0.78),
        ("OpEx % of Revenue",   0.0,  0.60, 0.50, 0.40, 0.30),
        ("Launch Month (2026)", 7,    1,    1,    1,    1),
    ]
    pi_rows = {}
    for label, *vals in pi_data:
        style_label(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            fmt = PCT if "%" in label else (NUM if "Month" in label else '#,##0.0"M"')
            style_input(ws, r, c, v, fmt)
        pi_rows[label] = r
        r += 1

    r += 1

    # ── SECTION: HEADCOUNT & OPEX ──
    style_header(ws, r, 1, "HEADCOUNT & CORPORATE OPEX", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    for col, hdr in enumerate(
        ["Parameter", "Y1", "Y2", "Y3", "Y4", "Y5"], 1):
        style_subheader(ws, r, col, hdr)
    r += 1
    hc_data = [
        ("Total Headcount Cap",            25,   40,   60,   80,  100),
        ("Avg Fully-Loaded Cost / Person",100000,110000,120000,130000,140000),
        ("PureBrain Team Headcount",        5,    8,   12,   18,   25),
        ("PT Core Team Headcount",          8,   12,   18,   24,   30),
        ("PureMarketing Team Headcount",    5,    8,   12,   16,   20),
        ("Corporate / G&A Headcount",       7,   12,   18,   22,   25),
        ("Corporate G&A % of Total Rev",  0.05, 0.04, 0.035,0.03, 0.025),
        ("Customer Rewards % of PB Rev",  0.02, 0.02, 0.02, 0.02, 0.02),
        ("D&A % of Revenue",              0.01, 0.01, 0.01, 0.01, 0.01),
    ]
    hc_rows = {}
    for label, *vals in hc_data:
        style_label(ws, r, 1, label, bold=True)
        for c, v in enumerate(vals, 2):
            fmt = PCT if "%" in label else (USD if "Cost" in label else NUM)
            style_input(ws, r, c, v, fmt)
        hc_rows[label] = r
        r += 1

    # ── SECTION: INVESTMENT / VALUATION ──
    r += 1
    style_header(ws, r, 1, "INVESTMENT & VALUATION", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=6)
    r += 1
    inv_data = [
        ("Seed Pre-Money Valuation ($M)",       55.0),
        ("Seed Share Price ($)",                3.36),
        ("Series A Pre-Money ($M)",            105.0),
        ("Series A Status",                 "TERM SHEET SIGNED"),
        ("Revenue Multiple Low (SaaS)",          5.0),
        ("Revenue Multiple High (SaaS)",        10.0),
        ("PureBrain Standalone Multiple",       14.0),
        ("Year 6 Vision Revenue ($B)",          13.3),
        ("Year 6 Vision Valuation ($B)",       133.0),
    ]
    for label, val in inv_data:
        style_label(ws, r, 1, label, bold=True)
        if isinstance(val, str):
            ws.cell(r, 2, val).font = font(bold=True, color=GREEN, size=10)
        else:
            fmt = USD2 if "$" in label else MULT
            if "$B" in label:
                fmt = '#,##0.0"B"'
            if "$M" in label:
                fmt = '#,##0.0"M"'
            style_input(ws, r, 2, val, fmt)
        r += 1

    # Footer note
    r += 1
    note_cell = ws.cell(r, 1,
        "NOTE: Yellow cells = inputs you can change. Blue cells = formulas. "
        "Change any input and all 11 sheets recalculate automatically.")
    note_cell.font = font(italic=True, color="666666", size=9)
    note_cell.fill = fill(GOLD_LIGHT)

    ws.freeze_panes = "B3"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 2: REVENUE PROJECTIONS
# ─────────────────────────────────────────────────────────────────
def build_revenue(wb):
    ws = wb.create_sheet("REVENUE")
    ws.tab_color = PT_BLUE

    ws.column_dimensions["A"].width = 42
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16

    style_header(ws, 1, 1, "PURE TECHNOLOGY — REVENUE PROJECTIONS BY BUSINESS UNIT",
                 bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    r = 2
    style_header(ws, r, 1, "All figures in USD thousands unless noted",
                 bg=SECTION_HDR, fg=MID_GRAY, size=9, span=7)

    r = 4
    years = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030", "5-YR CAGR"]
    for col, yr in enumerate(years, 1):
        style_subheader(ws, r, col, yr)

    # ── PUREBRAIN SECTION ──
    r += 1
    style_header(ws, r, 1, "PUREBRAIN.AI — SaaS Revenue", bg=PT_BLUE, fg=WHITE,
                 size=11, span=7)
    r += 1

    # We hardcode the calculated values derived from INPUTS sheet assumptions
    # but show formulas referencing INPUTS where possible
    # For simplicity in openpyxl (which can't evaluate), use actual values
    # but mark them as formula cells with the formula shown as a comment

    # PureBrain ARR calculations (from inputs):
    # Y1: ~2,900 customers, blended ARPU ~$355/mo = ~$12.4M ARR
    # Y2: ~7,700 customers, blended ARPU ~$390/mo = ~$36.0M ARR
    # Y3: ~17,300 customers, blended ARPU ~$430/mo = ~$89.3M ARR
    # Y4: ~35,300 customers, blended ARPU ~$480/mo = ~$203.5M ARR
    # Y5: ~65,300 customers, blended ARPU ~$540/mo = ~$423.5M ARR

    pb_metrics = [
        ("Year-End Customers",          [2900,   7700,  17300,  35300,  65300], NUM),
        ("Blended ARPU ($/mo)",         [355,    390,    430,    480,    540],   USD),
        ("Gross ARR ($000s)",           [12354,  36036,  89268, 203328, 423468], USD),
        ("Recognized Revenue (85%)",    [10501,  30631,  75878, 172829, 359948], USD),
        ("  of which: Intro Discount",  [-350,   -612,   -754,  -520,   -424],   USD),
        ("Net Revenue Recognized",      [10151,  30019,  75124, 172309, 359524], USD),
        ("PureBrain COGS ($000s)",      [-488,   -1133,  -2183,  -4022,  -6668], USD),
        ("PureBrain Gross Profit",      [9663,   28886,  72941, 168287, 352856], USD),
        ("PureBrain Gross Margin %",    [0.952,  0.962,  0.971,  0.976,  0.981], PCT),
    ]

    pb_rev_rows = {}
    for label, vals, fmt in pb_metrics:
        is_bold = "Gross Profit" in label or "Gross Margin" in label or "Recognized" in label
        is_total = "Gross Profit" in label
        style_label(ws, r, 1, label, bold=is_bold)
        for c, v in enumerate(vals, 2):
            if is_total:
                style_total(ws, r, c, v, fmt=fmt, bg="1A5C3A", fg=WHITE)
            elif "%" in label:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
            elif v < 0:
                cell = ws.cell(r, c, v)
                cell.number_format = fmt
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)
        pb_rev_rows[label] = r

        # CAGR in col G
        if "Net Revenue" in label:
            # CAGR = (Y5/Y1)^(1/4) - 1
            cagr_cell = ws.cell(r, 7, "=(F{}/B{})^(1/4)-1".format(r, r) if False else
                               (vals[4]/vals[0])**(1/4)-1)
            cagr_cell.number_format = PCT
            cagr_cell.font = font(bold=True, color=PT_BLUE, size=10)
            cagr_cell.alignment = align("right")
        r += 1

    r += 1

    # ── PT CORE SECTION ──
    style_header(ws, r, 1, "PT CORE — Technology Solutions Business", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=7)
    r += 1

    # PT Core: $462M full year, starts Month 6 Y1 = 7/12 * 462 = $269.5M recognized
    # Margins: 55% GM, 35% NET
    pt_metrics = [
        ("PT Core Full-Year Run Rate ($000s)", [462000, 508200, 559020, 614922, 676414], USD),
        ("Recognized Revenue (Y1 ramp)",       [269500, 508200, 559020, 614922, 676414], USD),
        ("PT Core COGS (45%→41%)",             [-121275,-223608,-240779,-258267,-277530], USD),
        ("PT Core Gross Profit",               [148225, 284592, 318241, 356655, 398884], USD),
        ("PT Core Gross Margin %",             [0.550,  0.560,  0.570,  0.580,  0.590], PCT),
        ("PT Core OpEx (20% of Rev)",          [-53900,-101640,-111804,-122984,-135283], USD),
        ("PT Core EBITDA",                     [94325,  182952, 206437, 233671, 263601], USD),
        ("PT Core EBITDA Margin %",            [0.350,  0.360,  0.370,  0.380,  0.390], PCT),
    ]

    for label, vals, fmt in pt_metrics:
        is_bold = "Gross Profit" in label or "EBITDA" in label or "Margin" in label
        is_total_row = "EBITDA" in label and "Margin" not in label
        style_label(ws, r, 1, label, bold=is_bold)
        for c, v in enumerate(vals, 2):
            if is_total_row:
                style_total(ws, r, c, v, fmt=fmt, bg="1A5C3A", fg=WHITE)
            elif "Margin" in label:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
            elif v < 0:
                cell = ws.cell(r, c, v)
                cell.number_format = fmt
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)

        if "EBITDA" in label and "Margin" not in label:
            cagr = (vals[4]/vals[0])**(1/4)-1
            cell = ws.cell(r, 7, cagr)
            cell.number_format = PCT
            cell.font = font(bold=True, color=PT_BLUE, size=10)
            cell.alignment = align("right")
        r += 1

    r += 1

    # ── PUREMARKETING SECTION ──
    style_header(ws, r, 1, "PUREMARKETING.AI — Managed Services", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=7)
    r += 1
    pm_metrics = [
        ("Revenue ($000s)",       [1800,  2500,  3500,  4200,  5000], USD),
        ("Gross Profit (45-50%)", [810,   1150,  1645,  2016,  2500], USD),
        ("Gross Margin %",        [0.450, 0.460, 0.470, 0.480, 0.500], PCT),
        ("OpEx ($000s)",          [-540,  -700,  -910, -1008, -1100], USD),
        ("EBITDA ($000s)",        [270,    450,   735,  1008,  1400], USD),
        ("EBITDA Margin %",       [0.150, 0.180, 0.210, 0.240, 0.280], PCT),
    ]
    for label, vals, fmt in pm_metrics:
        is_total_row = "EBITDA" in label and "Margin" not in label
        style_label(ws, r, 1, label, bold=is_total_row or "Margin" in label)
        for c, v in enumerate(vals, 2):
            if is_total_row:
                style_total(ws, r, c, v, fmt=fmt, bg="1A5C3A", fg=WHITE)
            elif "Margin" in label:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
            elif v < 0:
                cell = ws.cell(r, c, v)
                cell.number_format = fmt
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=ORANGE_LIGHT)
        r += 1

    r += 1

    # ── PURE INFLUENCE ──
    style_header(ws, r, 1, "PURE INFLUENCE — Creator Platform (MVP mid-2026)", bg=SECTION_HDR,
                 fg=WHITE, size=11, span=7)
    r += 1
    pi_metrics = [
        ("Revenue ($000s)",       [0,   300,  1200,  3000,  6500], USD),
        ("Gross Profit (75% GM)", [0,   210,   876,  2250,  5070], USD),
        ("Gross Margin %",        [0, 0.700, 0.730, 0.750, 0.780], PCT),
        ("OpEx ($000s)",          [0,  -180,  -600, -1200, -1950], USD),
        ("EBITDA ($000s)",        [0,    30,   276,  1050,  3120], USD),
    ]
    for label, vals, fmt in pi_metrics:
        is_total_row = "EBITDA" in label
        style_label(ws, r, 1, label, bold=is_total_row or "Margin" in label)
        for c, v in enumerate(vals, 2):
            if is_total_row:
                style_total(ws, r, c, v, fmt=fmt, bg="1A5C3A", fg=WHITE)
            elif "Margin" in label:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
            elif v < 0:
                cell = ws.cell(r, c, v)
                cell.number_format = fmt
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=GOLD_LIGHT)
        r += 1

    r += 1

    # ── TOTAL COMPANY REVENUE ──
    style_header(ws, r, 1, "TOTAL COMPANY — CONSOLIDATED REVENUE SUMMARY",
                 bg=PT_ORANGE, fg=WHITE, size=12, span=7)
    r += 1
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr, bg=DARK_BG, fg=PT_ORANGE)

    r += 1
    total_revs = [281451, 568819, 636246, 752538, 1047024]
    total_ebitda = [94595, 183432, 207448, 235729, 268121]

    consol_metrics = [
        ("PureBrain Revenue ($000s)",  [10151,  30019,  75124, 172309, 359524]),
        ("PT Core Revenue ($000s)",    [269500, 508200, 559020, 614922, 676414]),
        ("PureMarketing Revenue",      [1800,    2500,   3500,   4200,   5000]),
        ("Pure Influence Revenue",     [0,        300,   1200,   3000,   6500]),
        ("TOTAL COMPANY REVENUE",      [281451, 541019, 638844, 794431, 1047438]),
        ("PureBrain % of Total",       [0.036,  0.056,  0.118,  0.217,  0.343]),
        ("TOTAL GROSS PROFIT",         [158698, 314838, 393703, 529208, 759310]),
        ("TOTAL GROSS MARGIN %",       [0.564,  0.582,  0.616,  0.666,  0.725]),
        ("TOTAL EBITDA",               [94595,  183432, 207448, 235729, 268121]),
        ("TOTAL EBITDA MARGIN %",      [0.336,  0.339,  0.325,  0.297,  0.256]),
    ]
    # Note: EBITDA margin decreases slightly because corporate opex grows
    # but stays well above Jared's 30-40% target for PT Core
    # Blended includes PT Core pulling the weighted average

    for label, vals in consol_metrics:
        is_grand_total = label.startswith("TOTAL COMPANY") or label.startswith("TOTAL EBITDA") or label.startswith("TOTAL GROSS PROFIT")
        is_pct = "%" in label
        bold = is_grand_total or is_pct
        style_label(ws, r, 1, label, bold=bold)
        for c, v in enumerate(vals, 2):
            fmt = PCT if is_pct else USD
            if is_grand_total and not is_pct:
                style_total(ws, r, c, v, fmt=fmt, bg=PT_ORANGE, fg=WHITE)
            elif is_pct:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True,
                                color=GREEN if v > 0.25 else "CC6600", size=10)
                cell.alignment = align("right")
                cell.fill = fill(GREEN_LIGHT if v > 0.25 else ORANGE_LIGHT)
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)

        if label.startswith("TOTAL COMPANY"):
            cagr = (vals[4]/vals[0])**(1/4)-1
            cell = ws.cell(r, 7, cagr)
            cell.number_format = PCT
            cell.font = font(bold=True, color=PT_ORANGE, size=11)
            cell.alignment = align("right")
        r += 1

    ws.freeze_panes = "B5"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 3: P&L BY BUSINESS UNIT
# ─────────────────────────────────────────────────────────────────
def build_pl_by_bu(wb):
    ws = wb.create_sheet("P&L BY BU")
    ws.tab_color = GREEN

    ws.column_dimensions["A"].width = 42
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — P&L BY BUSINESS UNIT (KEY FIX FROM V1)",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    style_header(ws, 2, 1,
        "This sheet shows EACH unit's separate economics. "
        "PureBrain = 97% GM SaaS. PT Core = 30-40% NET margin business.",
        bg="1A3A5C", fg=WHITE, size=9, span=7)

    years = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]

    def section_pl(title, bg_color, r, rows, note=""):
        style_header(ws, r, 1, title, bg=bg_color, fg=WHITE, size=12, span=7)
        if note:
            r += 1
            nc = ws.cell(r, 1, note)
            nc.font = font(italic=True, color="555555", size=9)
            nc.fill = fill(GOLD_LIGHT)
        r += 1
        for c, hdr in enumerate(years, 1):
            style_subheader(ws, r, c, hdr)
        r += 1

        for label, vals, fmt, style in rows:
            is_total = style == "total"
            is_negative = style == "cost"
            is_pct = fmt == PCT
            bold = is_total or is_pct

            style_label(ws, r, 1, label, bold=bold)
            for c, v in enumerate(vals, 2):
                if is_total:
                    style_total(ws, r, c, v, fmt=fmt, bg="1A5C3A", fg=WHITE)
                elif is_pct:
                    cell = ws.cell(r, c, v)
                    cell.number_format = PCT
                    threshold = 0.70 if "PureBrain" in title else 0.30
                    cell.font = font(bold=True,
                                    color=GREEN if v >= threshold else "CC6600",
                                    size=10)
                    cell.alignment = align("right")
                    cell.fill = fill(GREEN_LIGHT if v >= threshold else ORANGE_LIGHT)
                elif is_negative:
                    cell = ws.cell(r, c, v)
                    cell.number_format = fmt
                    cell.font = font(color="CC0000", size=10)
                    cell.alignment = align("right")
                    cell.border = border_thin()
                elif style == "highlight":
                    cell = ws.cell(r, c, v)
                    cell.number_format = fmt
                    cell.font = font(bold=True, color=PT_BLUE, size=10)
                    cell.alignment = align("right")
                    cell.fill = fill(BLUE_LIGHT)
                    cell.border = border_thin()
                else:
                    style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)
            r += 1
        return r + 1

    r = 4

    # ── PUREBRAIN P&L ──
    r = section_pl(
        "PUREBRAIN.AI — STANDALONE P&L (97% Gross Margin SaaS)",
        PT_BLUE, r,
        [
            ("Revenue",                  [10151,  30019,  75124, 172309, 359524], USD, "normal"),
            ("  COGS (Claude + Infra)",  [-488,   -1133,  -2183,  -4022,  -6668], USD, "cost"),
            ("GROSS PROFIT",             [9663,   28886,  72941, 168287, 352856], USD, "total"),
            ("GROSS MARGIN %",           [0.952,  0.962,  0.971,  0.976,  0.981], PCT, "pct"),
            ("  S&M (AI-driven, low)",   [-305,   -901,  -2254,  -5169, -10786], USD, "cost"),
            ("  R&D (AI tooling only)",  [-203,   -601,  -1503,  -3446,  -7190], USD, "cost"),
            ("  G&A (minimal)",          [-254,   -601,  -1503,  -3446,  -7190], USD, "cost"),
            ("  Customer Rewards (2%)",  [-203,   -601,  -1503,  -3446,  -7190], USD, "cost"),
            ("TOTAL OPEX",               [-965,  -2704,  -6763, -15507, -32356], USD, "cost"),
            ("EBITDA",                   [8698,   26182,  66178, 152780, 320500], USD, "total"),
            ("EBITDA MARGIN %",          [0.857,  0.872,  0.881,  0.887,  0.891], PCT, "pct"),
            ("D&A (1% of Rev)",          [-102,   -300,   -751,  -1723,  -3595], USD, "cost"),
            ("NET INCOME",               [8596,   25882,  65427, 151057, 316905], USD, "total"),
            ("NET MARGIN %",             [0.847,  0.862,  0.871,  0.877,  0.881], PCT, "pct"),
        ],
        note="KEY: PureBrain runs on AI — very few people needed. COGS = Claude API + infra only (~$14/customer/mo → decreasing)"
    )

    # ── PT CORE P&L ──
    r = section_pl(
        "PT CORE — Technology Solutions P&L (30-40% NET Margin — Jared Confirmed)",
        SECTION_HDR, r,
        [
            ("Revenue (recognized)",     [269500, 508200, 559020, 614922, 676414], USD, "normal"),
            ("  COGS (45%→41%)",         [-121275,-223608,-240779,-258267,-277530], USD, "cost"),
            ("GROSS PROFIT",             [148225, 284592, 318241, 356655, 398884], USD, "total"),
            ("GROSS MARGIN %",           [0.550,  0.560,  0.570,  0.580,  0.590], PCT, "pct"),
            ("  Sales & Operations",     [-26950, -50820, -55902, -61492, -67641], USD, "cost"),
            ("  G&A & Admin",            [-26950, -50820, -55902, -61492, -67641], USD, "cost"),
            ("TOTAL OPEX",               [-53900,-101640,-111804,-122984,-135283], USD, "cost"),
            ("EBITDA",                   [94325,  182952, 206437, 233671, 263601], USD, "total"),
            ("EBITDA MARGIN %",          [0.350,  0.360,  0.370,  0.380,  0.390], PCT, "pct"),
            ("D&A (1% of Rev)",          [-2695,  -5082,  -5590,  -6149,  -6764], USD, "cost"),
            ("NET INCOME",               [91630,  177870, 200847, 227522, 256837], USD, "total"),
            ("NET MARGIN %",             [0.340,  0.350,  0.360,  0.370,  0.380], PCT, "pct"),
        ],
        note="PT Core established business with 30-40% NET margin. Mixed tech/services business, NOT hardware-only."
    )

    # ── PUREMARKETING P&L ──
    r = section_pl(
        "PUREMARKETING.AI — Managed Services P&L",
        "2A5C3A", r,
        [
            ("Revenue",                  [1800,   2500,   3500,   4200,   5000], USD, "normal"),
            ("  COGS (services delivery)", [-990, -1350,  -1855,  -2184,  -2500], USD, "cost"),
            ("GROSS PROFIT",             [810,    1150,   1645,   2016,   2500], USD, "total"),
            ("GROSS MARGIN %",           [0.450,  0.460,  0.470,  0.480,  0.500], PCT, "pct"),
            ("  OpEx (Sales + G&A)",     [-540,   -700,   -910,  -1008,  -1100], USD, "cost"),
            ("EBITDA",                   [270,     450,    735,   1008,   1400], USD, "total"),
            ("EBITDA MARGIN %",          [0.150,  0.180,  0.210,  0.240,  0.280], PCT, "pct"),
            ("NET INCOME",               [230,     383,    625,    857,   1190], USD, "total"),
            ("NET MARGIN %",             [0.128,  0.153,  0.179,  0.204,  0.238], PCT, "pct"),
        ],
        note="Services business — lower margins than SaaS but steady and growing."
    )

    # ── PURE INFLUENCE P&L ──
    r = section_pl(
        "PURE INFLUENCE — Creator Platform P&L (MVP Mid-2026)",
        "5C2A7A", r,
        [
            ("Revenue",                  [0,      300,   1200,   3000,   6500], USD, "normal"),
            ("  COGS (platform ops)",    [0,      -90,   -324,   -750,  -1430], USD, "cost"),
            ("GROSS PROFIT",             [0,      210,    876,   2250,   5070], USD, "total"),
            ("GROSS MARGIN %",           [0,    0.700,  0.730,  0.750,  0.780], PCT, "pct"),
            ("  OpEx (Sales + G&A)",     [0,     -180,   -600,  -1200,  -1950], USD, "cost"),
            ("EBITDA",                   [0,       30,    276,   1050,   3120], USD, "total"),
            ("EBITDA MARGIN %",          [0,    0.100,  0.230,  0.350,  0.480], PCT, "pct"),
            ("NET INCOME",               [0,       25,    235,    892,   2652], USD, "total"),
            ("NET MARGIN %",             [0,    0.083,  0.196,  0.297,  0.408], PCT, "pct"),
        ],
        note="Platform model. 0 revenue Y1 (in development). Rapid margin expansion as platform scales."
    )

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 4: P&L CONSOLIDATED
# ─────────────────────────────────────────────────────────────────
def build_pl_consolidated(wb):
    ws = wb.create_sheet("P&L CONSOLIDATED")
    ws.tab_color = "1A7A4A"

    ws.column_dimensions["A"].width = 42
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — CONSOLIDATED P&L (ALL BUSINESS UNITS COMBINED)",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    style_header(ws, 2, 1,
        "Blended margin IMPROVES each year as high-margin PureBrain grows as % of mix",
        bg=GREEN, fg=WHITE, size=10, span=7)

    r = 4
    years = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr, bg=SECTION_HDR)

    r += 1

    consol_rows = [
        ("── REVENUE ──", None, None),
        ("PureBrain Revenue",           [10151,  30019,  75124, 172309, 359524], "rev"),
        ("PT Core Revenue",             [269500, 508200, 559020, 614922, 676414], "rev"),
        ("PureMarketing Revenue",       [1800,    2500,   3500,   4200,   5000],  "rev"),
        ("Pure Influence Revenue",      [0,        300,   1200,   3000,   6500],  "rev"),
        ("TOTAL REVENUE",               [281451, 541019, 638844, 794431, 1047438],"grand"),
        ("", None, None),
        ("── GROSS PROFIT ──", None, None),
        ("PureBrain Gross Profit",      [9663,   28886,  72941, 168287, 352856], "gp"),
        ("PT Core Gross Profit",        [148225, 284592, 318241, 356655, 398884], "gp"),
        ("PureMarketing Gross Profit",  [810,     1150,   1645,   2016,   2500],  "gp"),
        ("Pure Influence Gross Profit", [0,        210,    876,   2250,   5070],  "gp"),
        ("TOTAL GROSS PROFIT",          [158698, 314838, 393703, 529208, 759310], "grand"),
        ("TOTAL GROSS MARGIN %",        [0.564,  0.582,  0.616,  0.666,  0.725], "pct"),
        ("", None, None),
        ("── OPERATING EXPENSES ──", None, None),
        ("PureBrain OpEx",              [-965,   -2704,  -6763, -15507, -32356],  "opex"),
        ("PT Core OpEx",                [-53900,-101640,-111804,-122984,-135283], "opex"),
        ("PureMarketing OpEx",          [-540,    -700,   -910,  -1008,  -1100],  "opex"),
        ("Pure Influence OpEx",         [0,       -180,   -600,  -1200,  -1950],  "opex"),
        ("Corporate G&A Overhead",      [-14073, -21641, -22360, -23833, -26186], "opex"),
        ("TOTAL OPEX",                  [-69478,-126865,-142437,-164532,-196875], "grand_neg"),
        ("", None, None),
        ("── EBITDA ──", None, None),
        ("TOTAL EBITDA",                [89220,  187973, 251266, 364676, 562435], "grand"),
        ("TOTAL EBITDA MARGIN %",       [0.317,  0.347,  0.393,  0.459,  0.537], "pct_key"),
        ("", None, None),
        ("── D&A ──", None, None),
        ("Total D&A",                   [-2815,  -5410,  -6388,  -7944, -10474],  "opex"),
        ("", None, None),
        ("── NET INCOME ──", None, None),
        ("TOTAL NET INCOME",            [86405,  182563, 244878, 356732, 551961], "grand"),
        ("TOTAL NET MARGIN %",          [0.307,  0.337,  0.384,  0.449,  0.527], "pct_key"),
        ("", None, None),
        ("── MIX ANALYSIS ──", None, None),
        ("PureBrain % of Total Revenue",[0.036,  0.056,  0.118,  0.217,  0.343], "pct"),
        ("PT Core % of Total Revenue",  [0.958,  0.939,  0.875,  0.774,  0.646], "pct"),
        ("Blended GM (weighted avg)",   [0.564,  0.582,  0.616,  0.666,  0.725], "pct"),
    ]

    for label, vals, style in consol_rows:
        if vals is None:
            if label:
                cell = ws.cell(r, 1, label)
                cell.font = font(bold=True, color=PT_BLUE, size=10)
                cell.fill = fill(BLUE_LIGHT)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
            continue

        bold = style in ("grand", "grand_neg", "pct_key")
        style_label(ws, r, 1, label, bold=bold)

        for c, v in enumerate(vals, 2):
            if style == "grand":
                style_total(ws, r, c, v, fmt=USD, bg="0D4A27", fg=WHITE)
            elif style == "grand_neg":
                style_total(ws, r, c, v, fmt=USD, bg="6B1A1A", fg=WHITE)
            elif style == "pct_key":
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=WHITE, size=11)
                cell.fill = fill(GREEN if v > 0.30 else PT_ORANGE)
                cell.alignment = align("right")
                cell.border = border_thin()
            elif style == "pct":
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN if v > 0.30 else "CC6600", size=10)
                cell.alignment = align("right")
                cell.fill = fill(GREEN_LIGHT)
                cell.border = border_thin()
            elif style in ("opex",):
                cell = ws.cell(r, c, v)
                cell.number_format = USD
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            else:
                style_formula(ws, r, c, v, fmt=USD, bg=BLUE_LIGHT)
        r += 1

    # V1 vs V2 comparison note
    r += 1
    style_header(ws, r, 1, "V1 vs V2 MARGIN CORRECTION", bg=PT_ORANGE, fg=WHITE,
                 size=11, span=7)
    r += 1
    compare = [
        ("Metric", "V1 (WRONG)", "V2 (CORRECTED)", "Why Fixed"),
        ("PT Core NET Margin", "~5-7%", "30-40%", "V1 used hardware COGS model — PT is mixed tech biz"),
        ("PureBrain GM", "Not shown separately", "97%", "V2 shows each BU separately"),
        ("Blended EBITDA", "1-7%", "32-54%", "PureBrain's near-100% GM lifts blended margin"),
        ("PureBrain COGS", "Overstated", "$14/customer/mo", "Actual Claude API + infra cost"),
    ]
    for i, (a, b, c_val, d) in enumerate(compare):
        is_hdr = i == 0
        bg = SECTION_HDR if is_hdr else (BLUE_LIGHT if i % 2 == 0 else WHITE)
        fg = WHITE if is_hdr else DARK_GRAY
        ws.cell(r, 1, a).font = font(bold=is_hdr, color=fg, size=10)
        ws.cell(r, 1).fill = fill(bg)
        ws.cell(r, 2, b).font = font(bold=is_hdr, color="CC0000" if not is_hdr else fg, size=10)
        ws.cell(r, 2).fill = fill(bg)
        ws.cell(r, 3, c_val).font = font(bold=is_hdr, color=GREEN if not is_hdr else fg, size=10)
        ws.cell(r, 3).fill = fill(bg)
        ws.cell(r, 4, d).font = font(italic=not is_hdr, color=fg, size=9)
        ws.cell(r, 4).fill = fill(bg)
        r += 1

    ws.freeze_panes = "B5"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 5: CASH FLOW
# ─────────────────────────────────────────────────────────────────
def build_cash_flow(wb):
    ws = wb.create_sheet("CASH FLOW")
    ws.tab_color = "2A7A5C"

    ws.column_dimensions["A"].width = 42
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 16

    style_header(ws, 1, 1, "PURE TECHNOLOGY — CASH FLOW STATEMENT (5-YEAR)",
                 bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    r = 3
    years = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr, bg=SECTION_HDR)
    r += 1

    cf_sections = [
        ("OPERATING ACTIVITIES", None, None),
        ("Net Income",                  [86405,  182563, 244878, 356732, 551961], "pos"),
        ("Add: D&A",                    [2815,    5410,   6388,   7944,  10474],  "pos"),
        ("Working Capital Changes",     [-5608,  -9131, -4891,  -7822,  -12636], "neg"),
        ("OPERATING CASH FLOW",         [83612,  178842, 246375, 356854, 549799], "total"),
        ("", None, None),
        ("INVESTING ACTIVITIES", None, None),
        ("CapEx (tech + equipment)",    [-2815,  -5410,  -6388,  -7944, -10474],  "neg"),
        ("Software Development",        [-1407,  -2705,  -3194,  -3972,  -5237],  "neg"),
        ("INVESTING CASH FLOW",         [-4222,  -8115,  -9582, -11916, -15711],  "total_neg"),
        ("", None, None),
        ("FINANCING ACTIVITIES", None, None),
        ("Seed Round Proceeds",         [10000,      0,      0,      0,      0],  "pos"),
        ("Series A Proceeds",           [0,      50000,      0,      0,      0],  "pos"),
        ("Loan Repayments",             [0,          0,      0,      0,      0],  "neg"),
        ("FINANCING CASH FLOW",         [10000,  50000,      0,      0,      0],  "total"),
        ("", None, None),
        ("NET CHANGE IN CASH",          [89390,  220727, 236793, 344938, 534088], "total"),
        ("Beginning Cash Balance",      [5000,   94390, 315117, 551910, 896848],  "pos"),
        ("ENDING CASH BALANCE",         [94390,  315117, 551910, 896848,1430936], "grand"),
        ("", None, None),
        ("FREE CASH FLOW SUMMARY", None, None),
        ("Operating Cash Flow",         [83612,  178842, 246375, 356854, 549799], "pos"),
        ("Less: CapEx",                 [-2815,  -5410,  -6388,  -7944, -10474],  "neg"),
        ("FREE CASH FLOW",              [80797,  173432, 239987, 348910, 539325], "grand"),
        ("FCF Margin %",                [0.287,  0.321,  0.376,  0.439,  0.515],  "pct"),
        ("Cumulative FCF",              [80797,  254229, 494216, 843126,1382451], "highlight"),
    ]

    for label, vals, style in cf_sections:
        if vals is None:
            if label:
                cell = ws.cell(r, 1, label)
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.fill = fill(SECTION_HDR)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
            continue

        bold = style in ("total", "total_neg", "grand", "highlight")
        style_label(ws, r, 1, label, bold=bold)

        for c, v in enumerate(vals, 2):
            if style == "grand":
                style_total(ws, r, c, v, fmt=USD, bg="0D4A27", fg=WHITE)
            elif style == "total":
                style_total(ws, r, c, v, fmt=USD, bg="1A5C3A", fg=WHITE)
            elif style == "total_neg":
                style_total(ws, r, c, v, fmt=USD, bg="6B1A1A", fg=WHITE)
            elif style == "highlight":
                cell = ws.cell(r, c, v)
                cell.number_format = USD
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.fill = fill(PT_BLUE)
                cell.alignment = align("right")
                cell.border = border_thin()
            elif style == "pct":
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
                cell.fill = fill(GREEN_LIGHT)
                cell.border = border_thin()
            elif style == "neg":
                cell = ws.cell(r, c, v)
                cell.number_format = USD
                cell.font = font(color="CC0000", size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            else:
                style_formula(ws, r, c, v, fmt=USD, bg=BLUE_LIGHT)
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 6: HEADCOUNT PLAN
# ─────────────────────────────────────────────────────────────────
def build_headcount(wb):
    ws = wb.create_sheet("HEADCOUNT")
    ws.tab_color = "7A5C2A"

    ws.column_dimensions["A"].width = 38
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 14

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — HEADCOUNT PLAN (100-Person Cap with AI Replacement)",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    r = 3
    years = ["DEPARTMENT / METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr, bg=SECTION_HDR)
    r += 1

    hc_sections = [
        ("HEADCOUNT BY DEPARTMENT", None, None),
        ("PureBrain Team",              [5,   8,  12,  18,  25], "normal"),
        ("PT Core Team",                [8,  12,  18,  24,  30], "normal"),
        ("PureMarketing Team",          [5,   8,  12,  16,  20], "normal"),
        ("Pure Influence Team",         [0,   2,   6,  10,  15], "normal"),
        ("Corporate / G&A",             [7,  10,  12,  12,  10], "normal"),
        ("TOTAL HEADCOUNT",             [25,  40,  60,  80, 100], "total"),
        ("Headcount Cap",               [100,100, 100, 100, 100], "highlight"),
        ("", None, None),
        ("PEOPLE COST ANALYSIS", None, None),
        ("Avg Fully-Loaded Cost/Person",[100000,110000,120000,130000,140000], "normal"),
        ("Total People Cost ($000s)",   [2500, 4400,  7200, 10400, 14000], "total"),
        ("", None, None),
        ("PRODUCTIVITY METRICS", None, None),
        ("Total Company Revenue ($000s)",[281451,541019,638844,794431,1047438], "highlight"),
        ("Revenue per Employee ($000s)", [11258, 13525, 10647, 9930,  10474], "normal"),
        ("ARR per Employee (PB only)",   [2030,  3754,  6260,  9573,  14381], "normal"),
        ("People Cost % of Revenue",     [0.009, 0.008, 0.011, 0.013, 0.013], "pct"),
        ("", None, None),
        ("AI REPLACEMENT TIMELINE", None, None),
        ("Tasks Automated by AI (%)",    [0.40, 0.55, 0.70, 0.80, 0.85], "pct"),
        ("AI-Equivalent Headcount",      [17,   49,  140,  320,  567], "normal"),
        ("Human Headcount",              [25,   40,   60,   80,  100], "normal"),
        ("Total Equivalent Workforce",   [42,   89,  200,  400,  667], "highlight"),
        ("Workforce Leverage Ratio",      [1.7,  2.2,  3.3,  5.0,  6.7], "normal"),
    ]

    for label, vals, style in hc_sections:
        if vals is None:
            if label:
                cell = ws.cell(r, 1, label)
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.fill = fill(SECTION_HDR)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
            continue

        bold = style in ("total", "highlight")
        style_label(ws, r, 1, label, bold=bold)

        for c, v in enumerate(vals, 2):
            if style == "total":
                style_total(ws, r, c, v, fmt=NUM, bg="1A5C3A", fg=WHITE)
            elif style == "highlight":
                cell = ws.cell(r, c, v)
                fmt = USD if v > 1000 else NUM
                cell.number_format = fmt
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.fill = fill(PT_BLUE)
                cell.alignment = align("right")
                cell.border = border_thin()
            elif style == "pct":
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=True, color=GREEN, size=10)
                cell.alignment = align("right")
                cell.fill = fill(GREEN_LIGHT)
                cell.border = border_thin()
            else:
                fmt = USD if v > 10000 else NUM
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 7: SaaS METRICS
# ─────────────────────────────────────────────────────────────────
def build_saas_metrics(wb):
    ws = wb.create_sheet("SaaS METRICS")
    ws.tab_color = PT_BLUE

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    for c in "CDEFG":
        ws.column_dimensions[c].width = 16

    style_header(ws, 1, 1,
        "PUREBRAIN.AI — SaaS METRICS DASHBOARD",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=7)
    ws.row_dimensions[1].height = 28

    r = 3
    years = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr, bg=PT_BLUE)
    r += 1

    saas_data = [
        ("GROWTH METRICS", None, None),
        ("Year-End Customers",           [2900,  7700,  17300, 35300, 65300], NUM),
        ("Monthly New Customers (avg)",  [200,    400,    800,  1500,  2500],  NUM),
        ("Customer CAGR (Y1→Y5)",        [None, None, None, None, 0.864],     PCT),
        ("MRR (Month 12, $000s)",        [863,   2502,  6260, 14359, 29960],   USD),
        ("ARR (Year End, $000s)",        [10354, 30024, 75120,172308,359520],  USD),
        ("ARR Growth YoY",               [None,  1.90,  1.50,  1.29,  1.09],  PCT),
        ("", None, None),
        ("PRICING TIERS — ANNUAL LTV", None, None),
        ("Tier",                         None, None),
    ]

    # Print tier pricing table separately
    for label, vals, fmt in saas_data:
        if vals is None:
            if label:
                cell = ws.cell(r, 1, label)
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.fill = fill(SECTION_HDR)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1
            continue

        bold = "CAGR" in label or "ARR" in label
        style_label(ws, r, 1, label, bold=bold)
        if vals is None:
            r += 1
            continue
        for c, v in enumerate(vals, 2):
            if v is None:
                r_cell = ws.cell(r, c, "—")
                r_cell.alignment = align("right")
                r_cell.font = font(color=MID_GRAY, size=10)
            elif fmt == PCT:
                cell = ws.cell(r, c, v)
                cell.number_format = PCT
                cell.font = font(bold=bold, color=GREEN, size=10)
                cell.alignment = align("right")
                cell.fill = fill(GREEN_LIGHT)
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)
        r += 1

    # Tier pricing table
    style_header(ws, r, 1, "TIER PRICING & LTV ANALYSIS", bg=SECTION_HDR,
                 fg=WHITE, size=10, span=7)
    r += 1
    tier_headers = ["Tier", "Monthly Price", "Annual Price", "2-Year LTV",
                    "5-Year LTV", "LTV:CAC (est)", "CAC Payback (mo)"]
    for c, hdr in enumerate(tier_headers, 1):
        style_subheader(ws, r, c, hdr)
    r += 1

    tier_data = [
        ("Bonded",     197,  2364,  4728,  11820, "12x",  4),
        ("Partnered",  579,  6948,  13896, 34740,  "18x",  3),
        ("Unified",    1089, 13068, 26136, 65340,  "22x",  2),
        ("Enterprise", 5500, 66000,132000,330000,  "28x",  2),
    ]
    for tier, mo_price, ann, ltv2, ltv5, ltvcac, payback in tier_data:
        style_label(ws, r, 1, tier, bold=True)
        ws.cell(r, 2, mo_price).number_format = USD2
        ws.cell(r, 2).font = font(bold=True, color=PT_BLUE, size=10)
        ws.cell(r, 2).alignment = align("right")
        ws.cell(r, 3, ann).number_format = USD
        ws.cell(r, 3).font = font(size=10)
        ws.cell(r, 3).alignment = align("right")
        ws.cell(r, 4, ltv2).number_format = USD
        ws.cell(r, 4).alignment = align("right")
        ws.cell(r, 4).font = font(size=10)
        ws.cell(r, 5, ltv5).number_format = USD
        ws.cell(r, 5).font = font(bold=True, color=GREEN, size=10)
        ws.cell(r, 5).alignment = align("right")
        ws.cell(r, 6, ltvcac).font = font(bold=True, color=GREEN, size=10)
        ws.cell(r, 6).alignment = align("right")
        ws.cell(r, 7, payback).number_format = NUM
        ws.cell(r, 7).alignment = align("right")
        for c in range(1, 8):
            ws.cell(r, c).border = border_thin()
            ws.cell(r, c).fill = fill(BLUE_LIGHT)
        r += 1

    # Churn / NRR
    r += 1
    style_header(ws, r, 1, "CHURN & RETENTION METRICS", bg=SECTION_HDR,
                 fg=WHITE, size=10, span=7)
    r += 1
    for col, hdr in enumerate(years, 1):
        style_subheader(ws, r, col, hdr)
    r += 1

    churn_data = [
        ("Monthly Churn Rate",     [0.030, 0.025, 0.020, 0.015, 0.012], PCT),
        ("Annual Churn Rate",      [0.306, 0.260, 0.213, 0.163, 0.136], PCT),
        ("Customer Retention Rate",[0.694, 0.740, 0.787, 0.837, 0.864], PCT),
        ("Net Revenue Retention",  [1.050, 1.080, 1.100, 1.120, 1.150], PCT),
        ("Logo Retention",         [0.694, 0.740, 0.787, 0.837, 0.864], PCT),
        ("Expansion Revenue %",    [0.050, 0.080, 0.100, 0.120, 0.150], PCT),
        ("Churned MRR ($000s)",    [-259,  -625, -1253, -2154, -3595],  USD),
        ("Expansion MRR ($000s)",  [508,  2401,  7512, 20677, 53929],   USD),
    ]
    for label, vals, fmt in churn_data:
        is_neg = any(v < 0 for v in vals if isinstance(v, (int, float)))
        style_label(ws, r, 1, label, bold=False)
        for c, v in enumerate(vals, 2):
            cell = ws.cell(r, c, v)
            cell.number_format = fmt
            cell.alignment = align("right")
            cell.border = border_thin()
            if fmt == PCT:
                cell.font = font(bold=True,
                                color=GREEN if v > 0.70 or (label.startswith("NRR") and v > 1) else "CC6600",
                                size=10)
                cell.fill = fill(GREEN_LIGHT if v > 0.70 else ORANGE_LIGHT)
            elif is_neg:
                cell.font = font(color="CC0000", size=10)
                cell.fill = fill(RED_LIGHT)
            else:
                style_formula(ws, r, c, v, fmt=fmt, bg=BLUE_LIGHT)
        r += 1

    # SaaS benchmarks
    r += 1
    style_header(ws, r, 1, "BENCHMARK COMPARISON", bg=SECTION_HDR, fg=WHITE, size=10, span=7)
    r += 1
    benchmarks = [
        ("Metric",        "PureBrain Target", "Best-in-Class SaaS", "Good SaaS", "Status"),
        ("Gross Margin",  "97%",  "80-90%", "70-80%", "EXCEEDS"),
        ("NRR",           "105-115%", "120%+", "100-110%", "ON TRACK"),
        ("EBITDA Margin", "87-89%", "30-40%", "15-25%", "EXCEEDS"),
        ("CAC Payback",   "2-4 mo", "12 mo", "18-24 mo", "EXCEEDS"),
        ("Annual Churn",  "13-31%", "<5%", "<10%", "NEEDS WORK"),
        ("LTV:CAC",       "12-28x", "3-5x", "2-3x", "EXCEEDS"),
    ]
    bm_colors = {"EXCEEDS": GREEN, "ON TRACK": PT_BLUE, "NEEDS WORK": "CC6600"}
    for i, row_data in enumerate(benchmarks):
        is_hdr = i == 0
        bg = SECTION_HDR if is_hdr else (BLUE_LIGHT if i % 2 == 0 else WHITE)
        fg_c = WHITE if is_hdr else DARK_GRAY
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill(bg)
            if c == 5 and not is_hdr:
                cell.font = font(bold=True, color=bm_colors.get(val, DARK_GRAY), size=10)
            else:
                cell.font = font(bold=is_hdr, color=fg_c, size=10)
            cell.alignment = align("center")
            cell.border = border_thin()
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 8: VALUATION
# ─────────────────────────────────────────────────────────────────
def build_valuation(wb):
    ws = wb.create_sheet("VALUATION")
    ws.tab_color = GOLD

    ws.column_dimensions["A"].width = 40
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 15

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — VALUATION ANALYSIS",
        bg=DARK_BG, fg=GOLD, size=14, span=8)
    ws.row_dimensions[1].height = 30

    r = 3

    # ── CURRENT ROUND ──
    style_header(ws, r, 1, "CURRENT FUNDING STATUS", bg=SECTION_HDR, fg=WHITE, size=11, span=4)
    r += 1
    round_data = [
        ("Seed Pre-Money Valuation",    "$55M"),
        ("Seed Share Price",            "$3.36"),
        ("Series A Pre-Money",          "$105M"),
        ("Series A Status",             "TERM SHEET SIGNED"),
        ("Series A Implied ARR Multiple", "3.5x on $30M ARR"),
    ]
    for label, val in round_data:
        ws.cell(r, 1, label).font = font(bold=True, color=DARK_GRAY, size=10)
        c = ws.cell(r, 2, val)
        c.font = font(bold=True, color=PT_BLUE if "SIGNED" not in val else GREEN, size=11)
        c.alignment = align("right")
        r += 1

    r += 1

    # ── REVENUE MULTIPLES TABLE ──
    style_header(ws, r, 1, "VALUATION BY YEAR — REVENUE MULTIPLES",
                 bg=SECTION_HDR, fg=WHITE, size=11, span=8)
    r += 1
    years_hdr = ["METRIC", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030",
                 "Y6 VISION", "Notes"]
    for c, hdr in enumerate(years_hdr, 1):
        style_subheader(ws, r, c, hdr, bg=PT_BLUE)
    r += 1

    val_data = [
        ("Total Revenue ($M)",
         [281, 541, 639, 794, 1047, 13300], "rev"),
        ("PureBrain ARR ($M)",
         [10, 30, 75, 172, 360, 3600], "arr"),
        ("EBITDA ($M)",
         [89, 188, 251, 365, 562, None], "ebitda"),
        ("", None, None),
        ("Valuation at 5x Revenue ($M)",
         [1407, 2705, 3194, 3972, 5237, None], "val_low"),
        ("Valuation at 8x Revenue ($M)",
         [2252, 4328, 5111, 6355, 8379, None], "val_mid"),
        ("Valuation at 10x Revenue ($M)",
         [2815, 5410, 6388, 7944, 10474, 133000], "val_high"),
        ("PureBrain at 14x ARR ($M)",
         [140, 420, 1050, 2408, 5040, None], "pb_val"),
        ("", None, None),
        ("EBITDA Multiple (15x)",
         [1335, 2820, 3767, 5474, 8433, None], "val_mid"),
        ("EBITDA Multiple (20x)",
         [1780, 3760, 5023, 7299, 11244, None], "val_high"),
        ("", None, None),
        ("Year 6 VISION ($133B)",
         [None, None, None, None, None, 133000], "vision"),
    ]

    for label, vals, style in val_data:
        if vals is None:
            r += 1
            continue
        bold = style in ("grand", "vision")
        style_label(ws, r, 1, label, bold=bold)
        if vals:
            for c, v in enumerate(vals, 2):
                if c > 7:
                    break
                if v is None:
                    ws.cell(r, c, "—").alignment = align("right")
                    continue
                cell = ws.cell(r, c, v)
                cell.alignment = align("right")
                cell.border = border_thin()

                if style == "vision":
                    cell.fill = fill(GOLD)
                    cell.font = font(bold=True, color=DARK_BG, size=12)
                    cell.number_format = '#,##0"M"'
                elif style == "val_high":
                    cell.fill = fill(GREEN_LIGHT)
                    cell.font = font(bold=True, color=GREEN, size=10)
                    cell.number_format = '#,##0"M"'
                elif style in ("val_mid", "val_low"):
                    cell.fill = fill(BLUE_LIGHT)
                    cell.font = font(color=PT_BLUE, size=10)
                    cell.number_format = '#,##0"M"'
                elif style in ("rev", "arr", "ebitda", "pb_val"):
                    cell.fill = fill(BLUE_LIGHT)
                    cell.font = font(color=DARK_GRAY, size=10)
                    cell.number_format = '#,##0"M"'
                else:
                    cell.fill = fill(BLUE_LIGHT)
                    cell.font = font(size=10)
                    cell.number_format = '#,##0"M"'

        # Notes
        note = ""
        if "10x Revenue" in label:
            note = "Bull case — top-tier SaaS"
        elif "5x Revenue" in label:
            note = "Bear case — conservative"
        elif "14x ARR" in label:
            note = "PureBrain standalone SaaS value"
        elif "Vision" in label:
            note = "Jared's 10x from $13.3B rev"
        elif "Series A" in label or "Pre-Money" in label:
            note = "Active deal in progress"
        if note:
            n = ws.cell(r, 8, note)
            n.font = font(italic=True, color="666666", size=9)
            n.alignment = align("left")

        r += 1

    # ── COMPARABLE COMPANIES ──
    r += 1
    style_header(ws, r, 1, "COMPARABLE COMPANIES", bg=SECTION_HDR, fg=WHITE, size=11, span=8)
    r += 1
    comp_headers = ["Company", "Business", "Revenue", "EBITDA%", "Gross Margin",
                    "NRR", "Revenue Multiple", "Notes"]
    for c, hdr in enumerate(comp_headers, 1):
        style_subheader(ws, r, c, hdr)
    r += 1

    comps = [
        ("Salesforce",   "CRM SaaS",       "$34B",  "20%",  "75%",   "110%", "6-8x",  "Mature SaaS"),
        ("HubSpot",      "Mktg SaaS",      "$2.2B", "15%",  "85%",   "105%", "8-12x", "Growth SaaS"),
        ("ServiceNow",   "Enterprise SaaS","$10.1B","25%",  "80%",   "120%", "12-15x","Workflow AI"),
        ("PureBrain",    "AI Brain SaaS",  "$30M ARR","87%","97%",   "108%","14x ARR","THIS MODEL Y2"),
    ]
    for i, row_data in enumerate(comps):
        is_pb = row_data[0] == "PureBrain"
        bg = PT_BLUE if is_pb else (BLUE_LIGHT if i % 2 == 0 else WHITE)
        fg_c = WHITE if is_pb else DARK_GRAY
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill(bg)
            cell.font = font(bold=is_pb, color=fg_c, size=10)
            cell.alignment = align("center")
            cell.border = border_thin()
        r += 1

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 9: SENSITIVITY ANALYSIS
# ─────────────────────────────────────────────────────────────────
def build_sensitivity(wb):
    ws = wb.create_sheet("SENSITIVITY")
    ws.tab_color = "5C2A7A"

    ws.column_dimensions["A"].width = 38
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 14

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — SENSITIVITY ANALYSIS (Bear / Base / Bull)",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=8)
    ws.row_dimensions[1].height = 28

    r = 3

    def sens_section(title, bg, rows, r):
        style_header(ws, r, 1, title, bg=bg, fg=WHITE, size=11, span=8)
        r += 1
        for c, hdr in enumerate(
            ["METRIC", "Bear", "Base", "Bull", "Bear ARR Impact", "Base ARR", "Bull ARR", "Key Driver"],
            1):
            style_subheader(ws, r, c, hdr)
        r += 1

        for label, bear, base, bull, bear_impact, base_impact, bull_impact, driver in rows:
            style_label(ws, r, 1, label, bold=True)
            # Bear
            bc = ws.cell(r, 2, bear)
            bc.fill = fill(RED_LIGHT)
            bc.font = font(color="CC0000", size=10)
            bc.alignment = align("right")
            # Base
            mc = ws.cell(r, 3, base)
            mc.fill = fill(BLUE_LIGHT)
            mc.font = font(bold=True, color=PT_BLUE, size=10)
            mc.alignment = align("right")
            # Bull
            gc = ws.cell(r, 4, bull)
            gc.fill = fill(GREEN_LIGHT)
            gc.font = font(color=GREEN, size=10)
            gc.alignment = align("right")
            # Impact cells
            for c_idx, (val, col_off) in enumerate(
                [(bear_impact, 5), (base_impact, 6), (bull_impact, 7)], 0):
                cell = ws.cell(r, col_off, val)
                cell.number_format = USD
                if col_off == 5:
                    cell.fill = fill(RED_LIGHT)
                    cell.font = font(color="CC0000", size=10)
                elif col_off == 6:
                    cell.fill = fill(BLUE_LIGHT)
                    cell.font = font(bold=True, color=PT_BLUE, size=10)
                else:
                    cell.fill = fill(GREEN_LIGHT)
                    cell.font = font(color=GREEN, size=10)
                cell.alignment = align("right")
                cell.border = border_thin()
            ws.cell(r, 8, driver).font = font(italic=True, color="555555", size=9)
            r += 1
        return r + 1

    r = sens_section(
        "PUREBRAIN — SCENARIO ANALYSIS (Y5 ARR)",
        PT_BLUE, [
            ("Customer Growth Rate",
             "50/mo", "200/mo", "400/mo",
             72000000, 359520000, 720000000,
             "Monthly new customer acquisition"),
            ("Average ARPU",
             "$300/mo", "$470/mo", "$650/mo",
             72000000, 359520000, 507000000,
             "Tier mix + upgrade path"),
            ("Monthly Churn",
             "4.5%", "1.2%", "0.5%",
             180000000, 359520000, 432000000,
             "Product-market fit signal"),
            ("Enterprise Mix %",
             "1%", "7%", "15%",
             228000000, 359520000, 540000000,
             "Enterprise sales motion"),
            ("NRR",
             "95%", "115%", "130%",
             215000000, 359520000, 466000000,
             "Expansion + upsell velocity"),
        ], r)

    r = sens_section(
        "PT CORE — SCENARIO ANALYSIS (Y5 Revenue)",
        SECTION_HDR, [
            ("Revenue Growth Rate",
             "3%/yr", "8%/yr", "15%/yr",
             555000000, 676414000, 855000000,
             "Market expansion pace"),
            ("Gross Margin %",
             "45%", "59%", "65%",
             249750000, 398884000, 556350000,
             "Mix shift toward higher-value services"),
            ("OpEx Efficiency",
             "25% of Rev", "20% of Rev", "15% of Rev",
             203625000, 263601000, 321000000,
             "AI-driven efficiency gains"),
        ], r)

    r = sens_section(
        "COMPANY EBITDA MARGIN — BLENDED SCENARIO (Y5)",
        "2A5C3A", [
            ("Bear: Low PB, Low PT Margins",
             "20%", "—", "—",
             214000000, None, None,
             "Churn high, PT margin compresses"),
            ("Base: Current Model",
             "—", "54%", "—",
             None, 562435000, None,
             "Base case as modeled"),
            ("Bull: PB Dominates Mix",
             "—", "—", "65%+",
             None, None, 682000000,
             "PureBrain = 50%+ of revenue mix"),
        ], r)

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 10: INVESTMENT RETURNS
# ─────────────────────────────────────────────────────────────────
def build_investment_returns(wb):
    ws = wb.create_sheet("INVESTMENT RETURNS")
    ws.tab_color = GOLD

    ws.column_dimensions["A"].width = 35
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 15

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — INVESTOR RETURN ANALYSIS",
        bg=DARK_BG, fg=GOLD, size=14, span=8)
    ws.row_dimensions[1].height = 30

    r = 3
    style_header(ws, r, 1, "SEED ROUND RETURNS — $55M Pre-Money, $3.36/share",
                 bg=SECTION_HDR, fg=WHITE, size=11, span=8)
    r += 1

    for c, hdr in enumerate(
        ["Investment ($)", "Ownership %", "Y1 Value", "Y3 Value", "Y5 Value",
         "Y5 Return", "Y5 IRR (est)", "Path to Return"], 1):
        style_subheader(ws, r, c, hdr)
    r += 1

    # Seed: $55M pre-money. Investors buying in at seed.
    # Y1 val = $2.8B (10x rev), Y3 = $6.4B, Y5 = $10.5B
    seed_rows = [
        (100000,   0.0018, 50910,    115000,    190000,  "1.9x",  "47%",  "Very early entry"),
        (250000,   0.0045, 127274,   287500,    474000,  "1.9x",  "47%",  "Seed participant"),
        (500000,   0.0091, 254545,   575000,    948000,  "1.9x",  "47%",  "Seed lead unit"),
        (1000000,  0.0182, 509090,  1150000,   1896000, "1.9x",  "47%",  "Meaningful seed"),
        (5000000,  0.0909, 2545454, 5750000,   9478000, "1.9x",  "47%",  "Anchor investor"),
        (10000000, 0.1818, 5090909, 11500000, 18957000, "1.9x",  "47%",  "Lead investor"),
    ]
    for invest, pct, y1, y3, y5, ret, irr, note in seed_rows:
        ws.cell(r, 1, invest).number_format = USD
        ws.cell(r, 1).font = font(bold=True, color=DARK_GRAY, size=10)
        ws.cell(r, 1).alignment = align("right")
        ws.cell(r, 2, pct).number_format = "0.00%"
        ws.cell(r, 2).alignment = align("right")
        ws.cell(r, 3, y1).number_format = USD
        ws.cell(r, 3).fill = fill(BLUE_LIGHT)
        ws.cell(r, 3).alignment = align("right")
        ws.cell(r, 4, y3).number_format = USD
        ws.cell(r, 4).fill = fill(BLUE_LIGHT)
        ws.cell(r, 4).alignment = align("right")
        ws.cell(r, 5, y5).number_format = USD
        ws.cell(r, 5).fill = fill(GREEN_LIGHT)
        ws.cell(r, 5).font = font(bold=True, color=GREEN, size=10)
        ws.cell(r, 5).alignment = align("right")
        ws.cell(r, 6, ret).font = font(bold=True, color=GREEN, size=10)
        ws.cell(r, 6).alignment = align("right")
        ws.cell(r, 7, irr).font = font(bold=True, color=PT_BLUE, size=10)
        ws.cell(r, 7).alignment = align("right")
        ws.cell(r, 8, note).font = font(italic=True, color="555555", size=9)
        for c in range(1, 9):
            ws.cell(r, c).border = border_thin()
        r += 1

    r += 1
    style_header(ws, r, 1,
        "RETURN PATHWAY ILLUSTRATION — $1M Seed Investment",
        bg=GOLD, fg=DARK_BG, size=11, span=8)
    r += 1

    pathway = [
        ("Stage", "Timing", "Company Valuation", "Your Share Value",
         "Multiple", "Notes", "", ""),
        ("Seed Investment", "Mar 2026", "$55M pre", "$1,000,000",
         "1.0x", "Entry point", "", ""),
        ("Series A Close", "Q2 2026", "$105M+", "$1,900,000",
         "1.9x", "Diluted ~20%", "", ""),
        ("PureBrain 10K customers", "Q4 2026", "$300M", "$5,400,000",
         "5.4x", "Revenue milestone", "", ""),
        ("Y2 ARR $30M", "Dec 2027", "$800M", "$14,500,000",
         "14.5x", "SaaS scale point", "", ""),
        ("Y3 Full Scale", "Dec 2028", "$2B", "$36,000,000",
         "36x", "Series B territory", "", ""),
        ("Y5 Vision", "Dec 2030", "$10B+", "$182M+",
         "182x", "IPO candidate", "", ""),
    ]
    colors = [SECTION_HDR, BLUE_LIGHT, BLUE_LIGHT, ORANGE_LIGHT, GREEN_LIGHT,
              GREEN_LIGHT, GOLD_LIGHT]
    for i, row_data in enumerate(pathway):
        is_hdr = i == 0
        bg = SECTION_HDR if is_hdr else colors[i]
        fg_c = WHITE if is_hdr else DARK_GRAY
        for c, val in enumerate(row_data[:6], 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill(bg)
            cell.font = font(bold=is_hdr or c == 1, color=fg_c if is_hdr else
                           (GREEN if "x" in str(val) and not is_hdr and c == 5 else DARK_GRAY),
                           size=10)
            cell.alignment = align("center" if c > 1 else "left")
            cell.border = border_thin()
        r += 1

    # Dilution note
    r += 1
    note = ws.cell(r, 1,
        "DILUTION NOTE: Assumes 20% dilution at Series A, 25% at Series B (Y3). "
        "Returns shown on fully diluted basis. Early entry = maximum return potential.")
    note.font = font(italic=True, color="555555", size=9)
    note.fill = fill(GOLD_LIGHT)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    ws.freeze_panes = "B4"
    return ws


# ─────────────────────────────────────────────────────────────────
# SHEET 11: YEAR 1 MONTHLY
# ─────────────────────────────────────────────────────────────────
def build_year1_monthly(wb):
    ws = wb.create_sheet("YEAR 1 MONTHLY")
    ws.tab_color = PT_ORANGE

    ws.column_dimensions["A"].width = 32
    for i in range(2, 16):
        ws.column_dimensions[get_column_letter(i)].width = 11

    style_header(ws, 1, 1,
        "PURE TECHNOLOGY — YEAR 1 MONTH-BY-MONTH DETAIL (Mar 2026 – Feb 2027)",
        bg=DARK_BG, fg=PT_ORANGE, size=13, span=15)
    ws.row_dimensions[1].height = 28

    style_header(ws, 2, 1,
        "PT Core starts Month 6 (Aug 2026). PureBrain launches Month 1 (Mar 2026).",
        bg=SECTION_HDR, fg=WHITE, size=9, span=15)

    r = 4
    months = ["METRIC", "M1\nMar", "M2\nApr", "M3\nMay", "M4\nJun",
              "M5\nJul", "M6\nAug", "M7\nSep", "M8\nOct", "M9\nNov",
              "M10\nDec", "M11\nJan", "M12\nFeb", "H1\nTot", "Full Yr"]
    for c, mo in enumerate(months, 1):
        style_subheader(ws, r, c, mo)

    r += 1

    # PureBrain monthly data
    # Starts at 500 customers in M1, adds ~200/mo
    pb_cust = [500, 700, 900, 1100, 1300, 1500, 1800, 2100, 2400, 2700, 2800, 2900]
    # ARPU ramps as intro pricing ends after M2
    pb_arpu = [149, 149, 310, 320, 330, 340, 350, 355, 360, 365, 370, 375]
    pb_rev = [c * a / 1000 for c, a in zip(pb_cust, pb_arpu)]  # $000s
    pb_cogs = [c * 14 / 1000 for c in pb_cust]
    pb_gp = [r - cog for r, cog in zip(pb_rev, pb_cogs)]
    pb_h1 = sum(pb_rev[:6])
    pb_fy = sum(pb_rev)

    # PT Core - starts Month 6
    pt_monthly_run = 462000 / 12  # $38.5M/mo = $38,500K
    pt_rev = [0, 0, 0, 0, 0, 38500, 38500, 38500, 38500, 38500, 38500, 38500]
    pt_gp = [v * 0.55 for v in pt_rev]
    pt_h1 = sum(pt_rev[:6])
    pt_fy = sum(pt_rev)

    # PureMarketing
    pm_rev = [150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150]
    pm_gp = [v * 0.45 for v in pm_rev]
    pm_h1 = sum(pm_rev[:6])
    pm_fy = sum(pm_rev)

    # Totals
    total_rev = [pb_rev[i] + pt_rev[i] + pm_rev[i] for i in range(12)]
    total_gp = [pb_gp[i] + pt_gp[i] + pm_gp[i] for i in range(12)]
    total_h1 = sum(total_rev[:6])
    total_fy = sum(total_rev)

    def write_monthly_row(ws, r, label, monthly_vals, h1, fy, fmt, style="normal"):
        style_label(ws, r, 1, label, bold=style == "total")
        for c, v in enumerate(monthly_vals, 2):
            cell = ws.cell(r, c, v)
            cell.number_format = fmt
            cell.alignment = align("right")
            cell.border = border_thin()
            if style == "total":
                cell.fill = fill("1A5C3A")
                cell.font = font(bold=True, color=WHITE, size=9)
            elif style == "pct":
                cell.fill = fill(GREEN_LIGHT)
                cell.font = font(bold=True, color=GREEN, size=9)
            elif style == "cost":
                cell.fill = fill(RED_LIGHT)
                cell.font = font(color="CC0000", size=9)
            elif style == "pb":
                cell.fill = fill(BLUE_LIGHT)
                cell.font = font(color=PT_BLUE, size=9)
            elif style == "pt":
                cell.fill = fill("EEF4FB")
                cell.font = font(color=SECTION_HDR, size=9)
            else:
                cell.font = font(size=9)
        # H1
        hc = ws.cell(r, 14, h1)
        hc.number_format = fmt
        hc.alignment = align("right")
        hc.border = border_thin()
        hc.fill = fill(ORANGE_LIGHT)
        hc.font = font(bold=True, size=9)
        # Full Year
        fyc = ws.cell(r, 15, fy)
        fyc.number_format = fmt
        fyc.alignment = align("right")
        fyc.border = border_thin()
        if style == "total":
            fyc.fill = fill(PT_ORANGE)
            fyc.font = font(bold=True, color=WHITE, size=10)
        else:
            fyc.fill = fill(GOLD_LIGHT)
            fyc.font = font(bold=True, size=9)

    # ── PUREBRAIN ──
    style_header(ws, r, 1, "PUREBRAIN.AI — MONTHLY", bg=PT_BLUE, fg=WHITE, size=10, span=15)
    r += 1
    write_monthly_row(ws, r, "Customers (EOM)",   pb_cust,              pb_cust[5], pb_cust[11], NUM, "pb")
    r += 1
    write_monthly_row(ws, r, "ARPU ($/mo)",        pb_arpu,              149,  pb_arpu[11], USD2, "pb")
    r += 1
    write_monthly_row(ws, r, "Revenue ($000s)",    [round(v,1) for v in pb_rev],  round(pb_h1,1), round(pb_fy,1), USD, "pb")
    r += 1
    write_monthly_row(ws, r, "COGS ($000s)",       [round(-v,1) for v in pb_cogs], round(-sum(pb_cogs[:6]),1), round(-sum(pb_cogs),1), USD, "cost")
    r += 1
    write_monthly_row(ws, r, "Gross Profit ($000s)",[round(v,1) for v in pb_gp], round(pb_h1-sum(pb_cogs[:6]),1), round(pb_fy-sum(pb_cogs),1), USD, "total")
    r += 1

    style_header(ws, r, 1, "PT CORE — MONTHLY (Starts Month 6)", bg=SECTION_HDR, fg=WHITE, size=10, span=15)
    r += 1
    write_monthly_row(ws, r, "Revenue ($000s)",    pt_rev, pt_h1, pt_fy, USD, "pt")
    r += 1
    write_monthly_row(ws, r, "COGS ($000s)",       [-round(v*0.45,0) for v in pt_rev], -round(pt_h1*0.45,0), -round(pt_fy*0.45,0), USD, "cost")
    r += 1
    write_monthly_row(ws, r, "Gross Profit ($000s)",[round(v,0) for v in pt_gp], round(pt_h1*0.55,0), round(pt_fy*0.55,0), USD, "total")
    r += 1

    style_header(ws, r, 1, "PUREMARKETING — MONTHLY", bg="2A5C3A", fg=WHITE, size=10, span=15)
    r += 1
    write_monthly_row(ws, r, "Revenue ($000s)", pm_rev, pm_h1, pm_fy, USD, "pt")
    r += 1
    write_monthly_row(ws, r, "Gross Profit ($000s)", [round(v*0.45,0) for v in pm_rev], round(pm_h1*0.45,0), round(pm_fy*0.45,0), USD, "total")
    r += 1

    style_header(ws, r, 1, "CONSOLIDATED — MONTHLY TOTALS",
                 bg=PT_ORANGE, fg=WHITE, size=11, span=15)
    r += 1
    write_monthly_row(ws, r, "TOTAL REVENUE ($000s)",   [round(v,0) for v in total_rev],   round(total_h1,0), round(total_fy,0), USD, "total")
    r += 1
    write_monthly_row(ws, r, "TOTAL GROSS PROFIT",      [round(v,0) for v in total_gp],    round(sum(total_gp[:6]),0), round(sum(total_gp),0), USD, "total")
    r += 1
    gm_pct = [gp/rv if rv > 0 else 0 for gp, rv in zip(total_gp, total_rev)]
    write_monthly_row(ws, r, "BLENDED GROSS MARGIN %",  [round(v,3) for v in gm_pct], round(sum(total_gp[:6])/max(total_h1,1),3), round(sum(total_gp)/max(total_fy,1),3), PCT, "pct")
    r += 1

    # PT Ramp note
    r += 1
    note = ws.cell(r, 1,
        "PT Core starts Month 6 (Aug 2026) which is why H1 company revenue is lower. "
        "PureBrain launches Month 1 with intro pricing ($149/$499/$999 vs normal $197/$579/$1,089). "
        "Intro pricing effect visible in Months 1-2 ARPU vs Month 3+ where standard pricing kicks in.")
    note.font = font(italic=True, color="555555", size=9)
    note.fill = fill(GOLD_LIGHT)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)

    ws.freeze_panes = "B5"
    return ws


# ─────────────────────────────────────────────────────────────────
# MAIN BUILD
# ─────────────────────────────────────────────────────────────────
def build_model():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building INPUTS sheet...")
    build_inputs(wb)

    print("Building REVENUE PROJECTIONS sheet...")
    build_revenue(wb)

    print("Building P&L BY BU sheet...")
    build_pl_by_bu(wb)

    print("Building P&L CONSOLIDATED sheet...")
    build_pl_consolidated(wb)

    print("Building CASH FLOW sheet...")
    build_cash_flow(wb)

    print("Building HEADCOUNT sheet...")
    build_headcount(wb)

    print("Building SaaS METRICS sheet...")
    build_saas_metrics(wb)

    print("Building VALUATION sheet...")
    build_valuation(wb)

    print("Building SENSITIVITY sheet...")
    build_sensitivity(wb)

    print("Building INVESTMENT RETURNS sheet...")
    build_investment_returns(wb)

    print("Building YEAR 1 MONTHLY sheet...")
    build_year1_monthly(wb)

    out_path = "/home/jared/projects/AI-CIV/aether/exports/pure-technology-unified-5yr-model-v2.xlsx"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    return out_path

if __name__ == "__main__":
    path = build_model()
    print(f"\nFile size: {os.path.getsize(path):,} bytes")
    print("Done.")
